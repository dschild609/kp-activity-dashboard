import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import io
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, KeepTogether, PageBreak
from reportlab.lib.enums import TA_CENTER
from reportlab.platypus import Image as RLImage

st.set_page_config(page_title="KP Staffing Dashboard", layout="wide", page_icon="📊")

# ── Password gate ──────────────────────────────────────────────────────────────
def _check_password():
    if st.session_state.get("authenticated"):
        return True
    st.image("kp_logo.png", width=180)
    st.markdown("## KP Staffing Activity Dashboard")
    st.markdown("Please enter the password to continue.")
    pw = st.text_input("Password", type="password", key="_login_pw")
    if st.button("Login"):
        correct = st.secrets.get("APP_PASSWORD", "kpstaffing2025")
        if pw == correct:
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    return False

if not _check_password():
    st.stop()
# ── End password gate ──────────────────────────────────────────────────────────

# KP Staffing brand colors
KP_CRIMSON  = "#940000"
KP_DARK     = "#7D1F32"
KP_LIGHT_BG = "#F2F4F6"
KP_TEXT     = "#131313"
KP_ACCENT   = "#E3F0FC"
KP_GRAY     = "#B2B5B8"

st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Poppins:wght@400;600;700&family=Roboto:wght@400;500&display=swap" rel="stylesheet">
<style>
    * { font-family: 'Roboto', sans-serif; }
    h1, h2, h3, .metric-value, .metric-label { font-family: 'Poppins', sans-serif; }

    .kp-header {
        display: flex; align-items: center; gap: 18px;
        padding: 18px 0 12px 0; border-bottom: 3px solid #940000; margin-bottom: 18px;
    }
    .kp-header img { height: 48px; }
    .kp-header-text h1 {
        margin: 0; font-family: 'Poppins', sans-serif; font-size: 1.5rem;
        font-weight: 700; color: #131313;
    }
    .kp-header-text p { margin: 2px 0 0 0; font-size: 0.85rem; color: #666; }

    .metric-card {
        background: white;
        border-radius: 10px;
        padding: 20px 14px;
        text-align: center;
        border: 1px solid #e0e0e0;
        border-top: 4px solid #940000;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
    .metric-value { font-size: 2.1rem; font-weight: 700; color: #940000; font-family: 'Poppins', sans-serif; }
    .metric-label { font-size: 0.72rem; color: #666; text-transform: uppercase; letter-spacing: 0.07em; margin-top: 5px; font-weight: 600; }
    .metric-sub { font-size: 0.75rem; color: #999; margin-top: 3px; }
    .metric-divider { border-top: 1px solid #e8e8e8; margin: 10px auto; width: 50%; }
    .metric-pct { font-size: 1.35rem; font-weight: 700; color: #7D1F32; margin-bottom: 2px; font-family: 'Poppins', sans-serif; }

    .section-header {
        font-size: 1rem; font-weight: 600; color: #940000;
        font-family: 'Poppins', sans-serif;
        margin: 24px 0 10px 0;
        border-bottom: 2px solid #940000;
        padding-bottom: 5px;
    }

    /* Global text */
    body, p, span, div, label, li { color: #131313 !important; }
    [data-testid="stAppViewContainer"] { background: #F2F4F6; }
    [data-testid="stMarkdownContainer"] p { color: #131313 !important; }

    /* Tabs */
    [data-testid="stTabs"] button { color: #131313 !important; font-weight: 600; }
    [data-testid="stTabs"] button[aria-selected="true"] { color: #940000 !important; border-bottom: 2px solid #940000; }

    /* Expander */
    [data-testid="stExpander"] summary { color: #131313 !important; font-weight: 600; }
    [data-testid="stExpander"] summary span { color: #131313 !important; }

    /* Dataframe wrapper */
    [data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }

    /* General Streamlit text overrides */
    .stMarkdown, .stText, .stCaption { color: #131313 !important; }
    h1, h2, h3, h4 { color: #131313 !important; }

    /* Sidebar */
    [data-testid="stSidebar"] { background: #8B1A1A; }
    [data-testid="stSidebar"] * { color: white !important; }
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] small,
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 { color: white !important; }
    [data-testid="stSidebar"] .stFileUploader { background: rgba(255,255,255,0.1); border-radius: 8px; padding: 8px; }
    [data-testid="stSidebar"] .stFileUploader * { color: white !important; }
    [data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.2); }

    /* Deploy button + sidebar collapse/expand arrows */
    [data-testid="stToolbar"] { color: white !important; }
    [data-testid="stToolbar"] button { color: white !important; }
    [data-testid="stToolbar"] svg { fill: white !important; stroke: white !important; }
    [data-testid="collapsedControl"] { background: rgba(255,255,255,0.25) !important; border: 1px solid rgba(255,255,255,0.4) !important; }
    [data-testid="collapsedControl"] svg { fill: white !important; stroke: white !important; color: white !important; }
    [data-testid="collapsedControl"] path { fill: white !important; stroke: white !important; }
    [data-testid="collapsedControl"] * { color: white !important; }
    [data-testid="stSidebar"] button[kind="header"] svg { fill: white !important; stroke: white !important; }
    [data-testid="baseButton-header"] { color: white !important; }
    [data-testid="baseButton-header"] svg { fill: white !important; stroke: white !important; }

    /* Streamlit toolbar & hamburger menu */
    header[data-testid="stHeader"] { background: #1a1a2e !important; }
    header[data-testid="stHeader"] * { color: white !important; }
    header[data-testid="stHeader"] svg { fill: white !important; stroke: white !important; }
    header[data-testid="stHeader"] button { color: white !important; }
    header[data-testid="stHeader"] path { fill: white !important; }
    /* Menu popup */
    div[data-baseweb="popover"] { background: #2c2c2c !important; }
    div[data-baseweb="popover"] * { color: white !important; }
    div[data-baseweb="popover"] li:hover { background: #444 !important; }
    div[data-baseweb="popover"] hr { border-color: #555 !important; }
    div[data-baseweb="menu"] { background: #2c2c2c !important; }
    div[data-baseweb="menu"] * { color: white !important; }
    div[data-baseweb="menu"] li:hover { background: #444 !important; }

    /* Generate PDF / Download buttons */
    [data-testid="stButton"] > button,
    [data-testid="stDownloadButton"] > button {
        background-color: #940000 !important;
        color: white !important;
        border: none !important;
        border-radius: 6px !important;
        padding: 10px 24px !important;
        font-family: 'Poppins', sans-serif !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        cursor: pointer !important;
        box-shadow: 0 2px 6px rgba(148,0,0,0.3) !important;
        transition: background-color 0.2s ease !important;
    }
    [data-testid="stButton"] > button p,
    [data-testid="stButton"] > button span,
    [data-testid="stDownloadButton"] > button p,
    [data-testid="stDownloadButton"] > button span {
        color: white !important;
    }
    [data-testid="stButton"] > button:hover,
    [data-testid="stDownloadButton"] > button:hover {
        background-color: #7D1F32 !important;
        color: white !important;
    }

    /* Delta comparison badges */
    .delta-good, .delta-good * { color: #2e7d32 !important; font-size: 1.05rem !important; font-weight: 700 !important; margin-top: 4px; }
    .delta-bad,  .delta-bad *  { color: #c62828 !important; font-size: 1.05rem !important; font-weight: 700 !important; margin-top: 4px; }
    .delta-neutral, .delta-neutral * { color: #888 !important; font-size: 1.05rem !important; font-weight: 600 !important; margin-top: 4px; }

    /* Comparison banner */
    .compare-banner {
        background: #f5f0f0; border-left: 4px solid #940000;
        border-radius: 6px; padding: 8px 14px; margin-bottom: 16px;
        font-size: 0.82rem; color: #131313;
    }

    /* Expander header — white background, dark text, red left border */
    div[data-testid="stExpander"] > details > summary {
        background-color: #ffffff !important;
        color: #131313 !important;
        border: 1px solid #e0e0e0 !important;
        border-left: 4px solid #940000 !important;
        border-radius: 6px !important;
        padding: 10px 14px !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
    }
    div[data-testid="stExpander"] > details > summary:hover {
        background-color: #fdf5f5 !important;
    }
    div[data-testid="stExpander"] > details > summary svg {
        fill: #940000 !important;
    }
</style>
""", unsafe_allow_html=True)

NO_START_REASONS = {"dropped out no start", "cancelled no start", "canceled no start"}
PROMO_REASONS    = {"changed title/promoted", "changed title / promoted", "changed positions", "changed title",
                    "converted to permanent"}
INVAL_STATUSES   = {"Ended by Employer", "Rejected"}
VOL_STATUS       = "Ended by Candidate"


def safe_int(val):
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def find_row(df, keyword, col=0):
    """Return index of first row where col contains keyword (case-insensitive)."""
    for i in range(len(df)):
        cell = str(df.iloc[i, col]) if pd.notna(df.iloc[i, col]) else ""
        if keyword.lower() in cell.lower():
            return i
    return None


def parse_report(file):
    raw1 = pd.read_excel(file, sheet_name=0, header=None)
    raw2 = pd.read_excel(file, sheet_name=1, header=None)

    # --- Meta ---
    title_cell = str(raw1.iloc[0, 0]) if pd.notna(raw1.iloc[0, 0]) else ""
    company = title_cell.split(" between ")[0].strip() if " between " in title_cell else title_cell
    date_range = title_cell.split(" between ")[-1].strip() if " between " in title_cell else ""

    # --- Summary counts: find row with "Total Placements" header, data is one row below ---
    summary_header_row = find_row(raw1, "Total Placements")
    if summary_header_row is not None and summary_header_row + 1 < len(raw1):
        hdr_vals = [str(v).strip() if pd.notna(v) else "" for v in raw1.iloc[summary_header_row]]
        data_row = raw1.iloc[summary_header_row + 1]
        total_placements_raw = safe_int(data_row[0])
        # Find "New Placements" column dynamically (old format: col 2; new format: col 1)
        try:
            new_col = hdr_vals.index("New Placements")
        except ValueError:
            new_col = 2  # fallback
        new_placements_raw = safe_int(data_row[new_col])
    else:
        total_placements_raw = 0
        new_placements_raw   = 0

    # --- Headcount: find "Period End Date" header, read date rows below until non-date ---
    period_row = find_row(raw1, "Period End Date")
    headcount_rows = []
    if period_row is not None:
        for i in range(period_row + 1, len(raw1)):
            row = raw1.iloc[i]
            cell = row[0]
            if pd.isna(cell):
                break
            try:
                dt = pd.to_datetime(cell)
                try:
                    hrs = float(row[2]) if pd.notna(row[2]) else 0.0
                except (ValueError, TypeError):
                    hrs = 0.0
                headcount_rows.append({
                    "Week Ending": dt,
                    "Headcount": safe_int(row[1]),
                    "Hours": hrs,
                })
            except Exception:
                break
    headcount_df = pd.DataFrame(headcount_rows)

    # --- Terms: find "End Reasons" label, then detect old vs new format ---
    end_reasons_row = find_row(raw1, "End Reasons")
    terms = []
    current_status = ""
    if end_reasons_row is not None:
        # Detect format: new format uses "Status: Ended by..." section headers
        new_terms_format = False
        for probe in range(end_reasons_row + 1, min(end_reasons_row + 6, len(raw1))):
            cell = str(raw1.iloc[probe, 0]).strip() if pd.notna(raw1.iloc[probe, 0]) else ""
            if cell.lower().startswith("status: ended"):
                new_terms_format = True
                break

        if new_terms_format:
            # New format: "Status: Ended by Candidate/Employer" section markers,
            # followed by a detail header row, then rows with:
            # col0=PlacementID, col1=Name, col2=SubDept, col3=JobTitle,
            # col4=StartDate, col5=ActualEndDate, col6=Days, col7=Status,
            # col8=EndReason, col9=Comments
            for i in range(end_reasons_row + 1, len(raw1)):
                row = raw1.iloc[i]
                cell0 = str(row[0]).strip() if pd.notna(row[0]) else ""
                cell1 = str(row[1]).strip() if pd.notna(row[1]) else ""

                if "ended by candidate" in cell0.lower():
                    current_status = VOL_STATUS
                    continue
                if "ended by employer" in cell0.lower():
                    current_status = "Ended by Employer"
                    continue
                # Skip column header rows and blank rows
                if cell1 in ("Name", "") or not cell1:
                    continue
                if not current_status:
                    continue

                name      = cell1
                reason    = str(row[8]).strip()  if len(row) > 8  and pd.notna(row[8])  else ""
                comment   = str(row[9]).strip()  if len(row) > 9  and pd.notna(row[9])  else ""
                s_dt      = pd.to_datetime(row[4], errors="coerce") if len(row) > 4 and pd.notna(row[4]) else None
                e_dt      = pd.to_datetime(row[5], errors="coerce") if len(row) > 5 and pd.notna(row[5]) else None
                staff_rep = str(row[11]).strip() if len(row) > 11 and pd.notna(row[11]) else ""
                job_title = str(row[3]).strip()  if len(row) > 3  and pd.notna(row[3])  else ""

                terms.append({
                    "Status":       current_status,
                    "Name":         name,
                    "End Reason":   reason,
                    "Count":        1,
                    "Comments":     comment,
                    "Start Date":   s_dt if s_dt is not None and not pd.isna(s_dt) else None,
                    "End Date":     e_dt if e_dt is not None and not pd.isna(e_dt) else None,
                    "Staffing Rep": staff_rep,
                    "Job Title":    job_title,
                })
        else:
            # Old format: Status in col0, Name col1, EndReason col2, Count col3, Comments col4
            terms_start = end_reasons_row + 2
            for i in range(terms_start, len(raw1)):
                row = raw1.iloc[i]
                s       = str(row[0]).strip() if pd.notna(row[0]) else ""
                name    = str(row[1]).strip() if pd.notna(row[1]) else ""
                reason  = str(row[2]).strip() if pd.notna(row[2]) else ""
                comment = str(row[4]).strip() if pd.notna(row[4]) else ""

                if s in INVAL_STATUSES or s == VOL_STATUS:
                    current_status = s
                    continue
                if not name or name == "Name":
                    continue
                raw_count = row[3]
                if pd.isna(raw_count):
                    continue
                try:
                    count = int(raw_count)
                except (ValueError, TypeError):
                    continue
                if current_status:
                    terms.append({
                        "Status":     current_status,
                        "Name":       name,
                        "End Reason": reason,
                        "Count":      count,
                        "Comments":   comment,
                    })

    terms_df = pd.DataFrame(terms) if terms else pd.DataFrame(
        columns=["Status", "Name", "End Reason", "Count", "Comments"])

    if not terms_df.empty:
        def classify(r):
            reason_lower = r["End Reason"].strip().lower()
            if reason_lower in ("layoff", "assignment complete", "assignment completed"):
                return "Layoff"
            if reason_lower in NO_START_REASONS:
                return "DONS"
            if any(p in reason_lower for p in PROMO_REASONS):
                return "Promoted"
            if r["Status"] in INVAL_STATUSES:
                return "Involuntary"
            return "Voluntary"
        terms_df["Type"] = terms_df.apply(classify, axis=1)
        # Strip DONS entirely — they never happened
        dons_count_raw = int(terms_df[terms_df["Type"] == "DONS"]["Count"].sum())
        terms_df = terms_df[terms_df["Type"] != "DONS"].reset_index(drop=True)
        # Strip Promoted — internal transfers, not true terms or new starts
        promoted_names = set(terms_df[terms_df["Type"] == "Promoted"]["Name"].tolist())
        terms_df = terms_df[terms_df["Type"] != "Promoted"].reset_index(drop=True)
    else:
        dons_count_raw = 0
        promoted_names = set()

    # --- Sheet 2: detect column positions from header row ---
    # Find the row containing "Start Date" in any column
    header2_row = None
    for r_idx in range(min(10, len(raw2))):
        row_vals = [str(v).strip() if pd.notna(v) else "" for v in raw2.iloc[r_idx]]
        if "Start Date" in row_vals:
            header2_row = r_idx
            break
    if header2_row is None:
        header2_row = 3  # fallback

    # Build header lookup: find ALL occurrences of each column name
    header_row_vals = [str(v).strip() if pd.notna(v) else "" for v in raw2.iloc[header2_row]]

    def _find_all(name):
        return [i for i, v in enumerate(header_row_vals) if v == name]

    # Detect if new format (has Job Title / Staffing Rep columns)
    has_jobtitle_cols = "Job Title" in header_row_vals
    has_staffrep_cols = "Staffing Rep" in header_row_vals

    # --- New Placements section (anchored by "Start Date") ---
    try:
        start_col = header_row_vals.index("Start Date")
    except ValueError:
        start_col = 4  # fallback
    # Find the "Name" column closest to (and before) Start Date
    name_candidates = [c for c in _find_all("Name") if c < start_col]
    name_col = name_candidates[-1] if name_candidates else 0
    # Pay Rate is always 1 col before date
    pay_col = start_col - 1
    # Job Title and Staffing Rep for new placements
    hire_jobtitle_col = None
    hire_staffrep_col = None
    if has_jobtitle_cols:
        jt_candidates = [c for c in _find_all("Job Title") if c < start_col]
        hire_jobtitle_col = jt_candidates[-1] if jt_candidates else None
    if has_staffrep_cols:
        sr_candidates = [c for c in _find_all("Staffing Rep") if c > start_col and c <= start_col + 2]
        hire_staffrep_col = sr_candidates[0] if sr_candidates else None

    # --- Ended Placements section (anchored by "End Date") ---
    try:
        end_col = header_row_vals.index("End Date")
    except ValueError:
        end_col = start_col + 4  # fallback
    end_name_candidates = [c for c in _find_all("Name") if c < end_col and c > start_col]
    end_name_col = end_name_candidates[-1] if end_name_candidates else end_col - 4
    end_pay_col = end_col - 1
    end_jobtitle_col = None
    end_staffrep_col = None
    if has_jobtitle_cols:
        jt_c = [c for c in _find_all("Job Title") if c < end_col and c > start_col]
        end_jobtitle_col = jt_c[-1] if jt_c else None
    if has_staffrep_cols:
        sr_c = [c for c in _find_all("Staffing Rep") if c > end_col and c <= end_col + 2]
        end_staffrep_col = sr_c[0] if sr_c else None

    # --- Converted Placements section (anchored by "Converted Date") ---
    try:
        conv_date_col = header_row_vals.index("Converted Date")
    except ValueError:
        conv_date_col = None
    conv_name_col = conv_pay_col = conv_jobtitle_col = conv_staffrep_col = None
    if conv_date_col is not None:
        cn_candidates = [c for c in _find_all("Name") if c < conv_date_col and c > start_col]
        conv_name_col = cn_candidates[-1] if cn_candidates else conv_date_col - 4
        conv_pay_col = conv_date_col - 1
        if has_jobtitle_cols:
            jt_c = [c for c in _find_all("Job Title") if c < conv_date_col and c > start_col]
            conv_jobtitle_col = jt_c[-1] if jt_c else None
        if has_staffrep_cols:
            sr_c = [c for c in _find_all("Staffing Rep") if c > conv_date_col and c <= conv_date_col + 2]
            conv_staffrep_col = sr_c[0] if sr_c else None

    hires, ended2, converted = [], [], []
    for i in range(header2_row + 1, len(raw2)):
        row = raw2.iloc[i]
        # New hires
        n = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
        s = pd.to_datetime(row[start_col], errors="coerce") if pd.notna(row[start_col]) else None
        if n and s and not pd.isna(s):
            rec = {"Name": n, "Pay Rate": row[pay_col] if pd.notna(row[pay_col]) else None, "Start Date": s}
            if hire_jobtitle_col is not None:
                rec["Job Title"] = str(row[hire_jobtitle_col]).strip() if pd.notna(row[hire_jobtitle_col]) else ""
            if hire_staffrep_col is not None:
                rec["Staffing Rep"] = str(row[hire_staffrep_col]).strip() if pd.notna(row[hire_staffrep_col]) else ""
            hires.append(rec)
        # Ended
        en = str(row[end_name_col]).strip() if pd.notna(row[end_name_col]) else ""
        ed = pd.to_datetime(row[end_col], errors="coerce") if pd.notna(row[end_col]) else None
        if en:
            rec2 = {"Name": en, "Pay Rate": row[end_pay_col] if pd.notna(row[end_pay_col]) else None, "End Date": ed if not pd.isna(ed) else None}
            if end_jobtitle_col is not None:
                rec2["Job Title"] = str(row[end_jobtitle_col]).strip() if pd.notna(row[end_jobtitle_col]) else ""
            if end_staffrep_col is not None:
                rec2["Staffing Rep"] = str(row[end_staffrep_col]).strip() if pd.notna(row[end_staffrep_col]) else ""
            ended2.append(rec2)
        # Converted
        if conv_name_col is not None:
            cn = str(row[conv_name_col]).strip() if pd.notna(row[conv_name_col]) else ""
            cd = pd.to_datetime(row[conv_date_col], errors="coerce") if pd.notna(row[conv_date_col]) else None
            if cn and cd and not pd.isna(cd):
                rec3 = {"Name": cn, "Pay Rate": row[conv_pay_col] if pd.notna(row[conv_pay_col]) else None, "Converted Date": cd}
                if conv_jobtitle_col is not None:
                    rec3["Job Title"] = str(row[conv_jobtitle_col]).strip() if pd.notna(row[conv_jobtitle_col]) else ""
                if conv_staffrep_col is not None:
                    rec3["Staffing Rep"] = str(row[conv_staffrep_col]).strip() if pd.notna(row[conv_staffrep_col]) else ""
                converted.append(rec3)

    hires_df     = pd.DataFrame(hires)
    ended2_df    = pd.DataFrame(ended2)
    converted_df = pd.DataFrame(converted)

    # Deduplicate each raw list: keep earliest start/converted date per person,
    # latest end date per person, so no one is counted twice
    if not hires_df.empty and "Start Date" in hires_df.columns:
        hires_df = (hires_df.sort_values("Start Date")
                            .drop_duplicates("Name", keep="first")
                            .reset_index(drop=True))
    if not ended2_df.empty and "End Date" in ended2_df.columns:
        ended2_df = (ended2_df.sort_values("End Date", na_position="first")
                              .drop_duplicates("Name", keep="last")
                              .reset_index(drop=True))
    if not converted_df.empty and "Converted Date" in converted_df.columns:
        converted_df = (converted_df.sort_values("Converted Date")
                                    .drop_duplicates("Name", keep="first")
                                    .reset_index(drop=True))

    # Remove promoted employees from new hires — internal transfers are not new starts
    if promoted_names and not hires_df.empty:
        removed_hires    = hires_df[hires_df["Name"].isin(promoted_names)]
        hires_df         = hires_df[~hires_df["Name"].isin(promoted_names)].reset_index(drop=True)
        new_placements_raw   = max(0, new_placements_raw   - len(removed_hires))
        total_placements_raw = max(0, total_placements_raw - len(removed_hires))

    # Fill in Staffing Rep / Job Title from terms lookup ONLY if not already present from Sheet 2
    if not hires_df.empty:
        has_sr = "Staffing Rep" in hires_df.columns and hires_df["Staffing Rep"].str.strip().ne("").any()
        has_jt = "Job Title" in hires_df.columns and hires_df["Job Title"].str.strip().ne("").any()
        if not has_sr or not has_jt:
            # Fall back to terms-based lookup for missing columns
            if not terms_df.empty and "Staffing Rep" in terms_df.columns:
                lookup_cols = [c for c in ["Name", "Staffing Rep", "Job Title"] if c in terms_df.columns]
                lookup = terms_df[lookup_cols].copy()
                for col in ["Staffing Rep", "Job Title"]:
                    if col in lookup.columns:
                        lookup = lookup[lookup[col].str.strip() != ""]
                lookup = lookup.drop_duplicates("Name")
                for col in ["Staffing Rep", "Job Title"]:
                    if col not in hires_df.columns or hires_df[col].str.strip().eq("").all():
                        if col in lookup.columns:
                            mapping = lookup.set_index("Name")[col]
                            hires_df[col] = hires_df["Name"].map(mapping).fillna("Unknown")
        # Ensure columns exist with defaults
        for col in ["Staffing Rep", "Job Title"]:
            if col not in hires_df.columns:
                hires_df[col] = "Unknown"
            else:
                hires_df[col] = hires_df[col].replace("", "Unknown").fillna("Unknown")

    # Merge start/end dates into terms — only if not already provided by new format
    if not terms_df.empty:
        if "Start Date" not in terms_df.columns and not hires_df.empty:
            terms_df = terms_df.merge(hires_df[["Name", "Start Date"]].drop_duplicates("Name"),
                                      on="Name", how="left")
        if "End Date" not in terms_df.columns and not ended2_df.empty:
            terms_df = terms_df.merge(ended2_df[["Name", "End Date"]].drop_duplicates("Name"),
                                      on="Name", how="left")

    # Deduplicate terms: one record per person — keep latest End Date so most recent
    # termination is used for retention calculations and detail tables
    if not terms_df.empty and "End Date" in terms_df.columns:
        terms_df = (terms_df.sort_values("End Date", na_position="first")
                            .drop_duplicates("Name", keep="last")
                            .reset_index(drop=True))

    # Derive converted_df from Page 1 End Reasons — "Converted to Permanent" is the
    # authoritative source; Sheet 2's converted column mixes in Changed Title/Promoted
    # entries and duplicates that inflate the count.
    if not terms_df.empty and "End Reason" in terms_df.columns:
        _conv_mask = terms_df["End Reason"].str.strip().str.lower() == "converted to permanent"
        if _conv_mask.any():
            _conv_cols = ["Name"] + [c for c in ["End Date", "Start Date", "Job Title", "Staffing Rep"] if c in terms_df.columns]
            converted_df = terms_df.loc[_conv_mask, _conv_cols].copy()
            converted_df = converted_df.rename(columns={"End Date": "Converted Date"})
            # Keep earliest converted date per person
            if "Converted Date" in converted_df.columns:
                converted_df = (converted_df.sort_values("Converted Date", na_position="last")
                                            .drop_duplicates("Name", keep="first")
                                            .reset_index(drop=True))
        # Remove converted employees from terms so they don't appear in turnover counts
        terms_df = terms_df[~_conv_mask].reset_index(drop=True)

    # --- Jobs section (optional — present in some report formats) ---
    jobs_row = find_row(raw1, "Jobs")
    jobs_df = pd.DataFrame()
    if jobs_row is not None:
        # Find the header row (contains "Open/Closed" or "Job Title")
        jobs_header_row = None
        for i in range(jobs_row, min(jobs_row + 5, len(raw1))):
            row_vals = [str(v).strip() if pd.notna(v) else "" for v in raw1.iloc[i]]
            if "Job Title" in row_vals or "Open/Closed" in row_vals:
                jobs_header_row = i
                break
        if jobs_header_row is not None:
            jobs_rows = []
            last_status = ""
            for i in range(jobs_header_row + 1, len(raw1)):
                row = raw1.iloc[i]
                # Stop if we hit another section
                cell0 = str(row[0]).strip() if pd.notna(row[0]) else ""
                if cell0 and cell0 not in ["Closed", "Open"] and not str(row[1]).strip().lstrip("0123456789").replace(" ", "") == "":
                    try:
                        int(str(row[1]).strip())  # Job ID is numeric
                    except (ValueError, TypeError):
                        if cell0 not in ["", "nan"]:
                            break
                if cell0 in ["Closed", "Open"]:
                    last_status = cell0
                job_id   = str(row[1]).strip() if pd.notna(row[1]) else ""
                title    = str(row[2]).strip() if pd.notna(row[2]) else ""
                if not job_id:
                    continue
                try:
                    int(job_id)
                except ValueError:
                    continue
                jobs_rows.append({
                    "Status":         last_status if last_status else cell0,
                    "Job ID":         job_id,
                    "Job Title":      title,
                    "Date Added":     pd.to_datetime(row[3], errors="coerce") if pd.notna(row[3]) else None,
                    "Date Closed":    pd.to_datetime(row[4], errors="coerce") if pd.notna(row[4]) else None,
                    "Days Opened":    str(row[5]).replace(" days","").replace(" day","").strip() if pd.notna(row[5]) else "0",
                    "Job Status":     str(row[6]).strip() if pd.notna(row[6]) else "",
                    "# Openings":     safe_int(row[7]),
                    "Placements":     safe_int(row[8]),
                    "Fill Rate":      float(row[9]) if pd.notna(row[9]) else 0.0,
                })
            jobs_df = pd.DataFrame(jobs_rows)
            if not jobs_df.empty:
                jobs_df["Days Opened"] = pd.to_numeric(jobs_df["Days Opened"], errors="coerce").fillna(0).astype(int)

    return {
        "company": company,
        "date_range": date_range,
        "total_placements_raw": total_placements_raw - dons_count_raw,
        "new_placements_raw":   new_placements_raw   - dons_count_raw,
        "headcount": headcount_df,
        "terms": terms_df,
        "hires": hires_df,
        "converted": converted_df,
        "jobs": jobs_df,
    }


def filter_data_by_dates(data, start_date, end_date):
    """Return a copy of data filtered to the given date window."""
    import copy
    d = copy.copy(data)

    # Filter headcount rows to weeks ending within window
    hc = data["headcount"].copy()
    if not hc.empty and "Week Ending" in hc.columns:
        hc = hc[(hc["Week Ending"] >= pd.Timestamp(start_date)) &
                (hc["Week Ending"] <= pd.Timestamp(end_date))]
    d["headcount"] = hc

    # Filter terms by End Date
    terms = data["terms"].copy()
    if not terms.empty and "End Date" in terms.columns:
        terms = terms[(terms["End Date"] >= pd.Timestamp(start_date)) &
                      (terms["End Date"] <= pd.Timestamp(end_date))]
    d["terms"] = terms

    # Filter hires by Start Date
    hires = data["hires"].copy()
    if not hires.empty and "Start Date" in hires.columns:
        hires = hires[(hires["Start Date"] >= pd.Timestamp(start_date)) &
                      (hires["Start Date"] <= pd.Timestamp(end_date))]
    d["hires"] = hires

    # Filter converted by Converted Date
    converted = data["converted"].copy()
    if not converted.empty and "Converted Date" in converted.columns:
        converted = converted[(converted["Converted Date"] >= pd.Timestamp(start_date)) &
                              (converted["Converted Date"] <= pd.Timestamp(end_date))]
    d["converted"] = converted

    return d


def compute_metrics(data):
    terms  = data["terms"]
    hires  = data["hires"]
    hc     = data["headcount"]
    empty  = pd.DataFrame()

    involuntary = terms[terms["Type"] == "Involuntary"] if not terms.empty else empty
    voluntary   = terms[terms["Type"] == "Voluntary"]   if not terms.empty else empty
    layoffs     = terms[terms["Type"] == "Layoff"]      if not terms.empty else empty

    inval_count  = int(involuntary["Count"].sum()) if not involuntary.empty else 0
    vol_count    = int(voluntary["Count"].sum())   if not voluntary.empty else 0
    layoff_count = int(layoffs["Count"].sum())     if not layoffs.empty else 0

    # --- Starting headcount calculation ---
    # Take the first week's on-site headcount, then subtract any new placements
    # whose start date falls on or before that first week-end date.
    # This isolates the employees who were already on site before the period began.
    starting_headcount = None
    first_week_end     = None
    new_hires_wk1      = 0

    if not hc.empty:
        first_week_end      = hc["Week Ending"].iloc[0]
        first_week_headcount = int(hc["Headcount"].iloc[0])

        if not hires.empty and "Start Date" in hires.columns:
            new_hires_wk1 = int(
                hires[hires["Start Date"] <= first_week_end].shape[0]
            )

        starting_headcount = max(first_week_headcount - new_hires_wk1, 1)

    # Total at-risk = starting headcount + total starts (DONS already stripped from both)
    total_starts = data["new_placements_raw"]
    if starting_headcount:
        at_risk = starting_headcount + total_starts
    else:
        at_risk = max(data["total_placements_raw"] - layoff_count, 1)

    total_hours = float(hc["Hours"].sum()) if not hc.empty and "Hours" in hc.columns else 0.0

    # --- Retention Tiers (7 / 30 / 60 day) — cascading, eligibility-aware ---
    total_new_starts = len(hires) if not hires.empty else 0
    retention_tiers = {}

    if total_new_starts > 0:
        # Cutoff = latest date in the dataset (determines who is evaluable)
        _cutoff_dates = []
        if not hc.empty and "Week Ending" in hc.columns:
            _cutoff_dates += hc["Week Ending"].dropna().tolist()
        if not terms.empty and "End Date" in terms.columns:
            _cutoff_dates += terms["End Date"].dropna().tolist()
        if not hires.empty and "Start Date" in hires.columns:
            _cutoff_dates += hires["Start Date"].dropna().tolist()
        _cutoff = max(_cutoff_dates) if _cutoff_dates else pd.Timestamp.now()

        # Build per-hire df with end date (if terminated)
        hire_df = hires[["Name", "Start Date"]].drop_duplicates("Name").copy()
        if not terms.empty and "Name" in terms.columns and "End Date" in terms.columns:
            _term_lookup = terms[["Name", "End Date", "End Reason"]].drop_duplicates("Name").copy()
            hire_df = hire_df.merge(_term_lookup, on="Name", how="left")
        else:
            hire_df["End Date"] = pd.NaT
            hire_df["End Reason"] = ""
        hire_df["Days Employed"] = (hire_df["End Date"] - hire_df["Start Date"]).dt.days

        _pool = hire_df.copy()  # cascades: each tier only evaluates survivors of prior tier

        for days in (7, 30, 60):
            eligible = _pool[
                _pool["Start Date"] + pd.Timedelta(days=days) <= pd.Timestamp(_cutoff)
            ].copy()

            if eligible.empty:
                retention_tiers[days] = {
                    "early_terms": pd.DataFrame(),
                    "early_count": 0,
                    "retained": 0,
                    "eligible": 0,
                    "pct": None,  # None = insufficient data
                }
                _pool = pd.DataFrame(columns=hire_df.columns)
                continue

            _passed_mask = eligible["End Date"].isna() | (eligible["Days Employed"] >= days)
            _passed = eligible[_passed_mask]
            _failed = eligible[~_passed_mask]

            # Early terms detail — pull full row from terms for reason/dates
            if not _failed.empty and not terms.empty:
                _et = terms[terms["Name"].isin(_failed["Name"])].drop_duplicates("Name").copy()
                if not _et.empty and "Start Date" in _et.columns and "End Date" in _et.columns:
                    _et = _et.dropna(subset=["Start Date", "End Date"])
                    if not _et.empty:
                        _et["Days Employed"] = (_et["End Date"] - _et["Start Date"]).dt.days
                else:
                    _et = pd.DataFrame()
            else:
                _et = pd.DataFrame()

            retention_tiers[days] = {
                "early_terms": _et,
                "early_count": len(_failed),
                "retained":    len(_passed),
                "eligible":    len(eligible),
                "pct":         len(_passed) / len(eligible) * 100,
            }
            _pool = _passed  # cascade survivors to next tier

    return {
        "adj_placements":      data["total_placements_raw"] - layoff_count,
        "starting_headcount":  starting_headcount,
        "first_week_end":      first_week_end,
        "new_hires_wk1":       new_hires_wk1,
        "at_risk":             at_risk,
        "inval_count":         inval_count,
        "vol_count":           vol_count,
        "layoff_count":        layoff_count,
        "total_hours":         total_hours,
        "inval_pct":           (inval_count / at_risk * 100) if at_risk > 0 else 0.0,
        "vol_pct":             (vol_count   / at_risk * 100) if at_risk > 0 else 0.0,
        "total_turnover_pct":  ((inval_count + vol_count) / at_risk * 100) if at_risk > 0 else 0.0,
        "involuntary":      involuntary,
        "voluntary":        voluntary,
        "layoffs":          layoffs,
        "total_new_starts": total_new_starts,
        "retention_tiers":  retention_tiers,
    }


def make_fig_headcount(headcount_df, include_hours=True):
    hc = headcount_df.copy()
    hc["Week"] = hc["Week Ending"].dt.strftime("%b %d")
    has_hours = include_hours and "Hours" in hc.columns and float(hc["Hours"].sum()) > 0
    max_hc = float(hc["Headcount"].max()) if not hc.empty else 10.0

    fig = go.Figure()

    # Headcount trace (left y-axis)
    fig.add_trace(go.Scatter(
        x=hc["Week"], y=hc["Headcount"],
        name="Headcount",
        mode="lines+markers",
        line=dict(color="#940000", width=3),
        marker=dict(size=10, color="#940000"),
        fill="tozeroy", fillcolor="rgba(148,0,0,0.07)",
        yaxis="y1",
    ))
    num_pts = len(hc)
    for idx, (_, row) in enumerate(hc.iterrows()):
        xs = 20 if idx == 0 else (-20 if idx == num_pts - 1 else 0)
        fig.add_annotation(
            x=row["Week"], y=float(row["Headcount"]),
            text=str(int(row["Headcount"])),
            showarrow=False, yshift=14, xshift=xs,
            font=dict(color="#940000", size=12),
            xref="x", yref="y",
        )

    # Hours trace (right y-axis)
    if has_hours:
        max_hrs = float(hc["Hours"].max())
        fig.add_trace(go.Scatter(
            x=hc["Week"], y=hc["Hours"],
            name="Hours",
            mode="lines+markers",
            line=dict(color="#1a6ea8", width=2, dash="dot"),
            marker=dict(size=8, color="#1a6ea8"),
            yaxis="y2",
        ))
        for idx2, (_, row) in enumerate(hc.iterrows()):
            xs2 = 20 if idx2 == 0 else (-20 if idx2 == num_pts - 1 else 0)
            fig.add_annotation(
                x=row["Week"], y=float(row["Hours"]),
                text=f"{int(row['Hours']):,}",
                showarrow=False, yshift=-18, xshift=xs2,
                font=dict(color="#1a6ea8", size=11),
                xref="x", yref="y2",
            )

    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=50, r=70, t=40, b=10),
        xaxis=dict(**DARK_AXIS),
        yaxis=dict(**DARK_AXIS, title=dict(text="Headcount", font=dict(color="#940000")), range=[0, max_hc * 1.30]),
        height=280,
        legend=dict(orientation="h", x=0.01, y=1.08, font=dict(size=11)),
        showlegend=bool(has_hours),
    )
    if has_hours:
        fig.update_layout(yaxis2=dict(
            title=dict(text="Hours", font=dict(color="#1a6ea8")),
            overlaying="y", side="right",
            showgrid=False,
            tickfont=dict(color="#1a6ea8"),
            range=[0, max_hrs * 1.30],
        ))
    return fig


def make_fig_donut(m):
    df = pd.DataFrame({
        "Type":  ["Involuntary", "Voluntary", "Layoff / Assignment Complete"],
        "Count": [m["inval_count"], m["vol_count"], m["layoff_count"]]
    })
    df = df[df["Count"] > 0]
    fig = px.pie(df, names="Type", values="Count",
                 color_discrete_sequence=["#940000", "#2C3E50", "#B2B5B8"],
                 hole=0.55)
    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=10, r=10, t=10, b=10),
        legend=dict(orientation="h", yanchor="bottom", y=-0.25,
                    font=dict(color="#131313", size=12)),
        height=260,
    )
    fig.update_traces(textinfo="percent+value")
    return fig


DARK_AXIS = dict(
    gridcolor="#e8e8e8",
    linecolor="#cccccc",
    tickfont=dict(color="#131313", size=12),
    title_font=dict(color="#131313", size=12),
)

def make_fig_reasons(df, color_seq, df_past=None):
    if df.empty:
        return go.Figure()
    filtered = df[df["End Reason"].str.strip() != ""]
    if filtered.empty:
        return go.Figure()
    curr = filtered.groupby("End Reason")["Count"].sum().reset_index(name="Current")

    if df_past is not None and not df_past.empty:
        past_f = df_past[df_past["End Reason"].str.strip() != ""]
        past  = past_f.groupby("End Reason")["Count"].sum().reset_index(name="Prior")
        merged = curr.merge(past, on="End Reason", how="outer").fillna(0)
        merged = merged.sort_values("Current", ascending=True)
        _cm = merged["Current"].max()
        _pm = merged["Prior"].max()
        max_count = int(max(_cm if pd.notna(_cm) else 0, _pm if pd.notna(_pm) else 0))
        if max_count == 0:
            return go.Figure()
        tick_step = max(1, round(max_count / 10))
        num_reasons = len(merged)
        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=merged["End Reason"], x=merged["Prior"],
            name="Prior Period", orientation="h",
            marker_color="rgba(180,180,180,0.6)",
            text=merged["Prior"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
        ))
        fig.add_trace(go.Bar(
            y=merged["End Reason"], x=merged["Current"],
            name="Current Period", orientation="h",
            marker_color=color_seq[0],
            text=merged["Current"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
        ))
        fig.update_layout(
            barmode="group",
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(color="#131313", size=12),
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(**DARK_AXIS, dtick=tick_step, range=[0, max_count * 1.25]),
            yaxis=dict(**DARK_AXIS),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(color="#131313")),
            height=max(300, num_reasons * 52),
        )
    else:
        grouped = curr.sort_values("Current", ascending=True)
        _cm = grouped["Current"].max()
        max_count = int(_cm) if pd.notna(_cm) and _cm > 0 else 0
        if max_count == 0:
            return go.Figure()
        tick_step = max(1, round(max_count / 10))
        num_reasons = len(grouped)
        fig = px.bar(grouped, x="Current", y="End Reason", orientation="h",
                     color="End Reason", color_discrete_sequence=color_seq)
        fig.update_layout(
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(color="#131313", size=12),
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(**DARK_AXIS, dtick=tick_step, range=[0, max_count * 1.15]),
            yaxis=dict(**DARK_AXIS),
            showlegend=False,
            height=max(300, num_reasons * 38),
        )
    return fig


def make_fig_jobs_fill(jobs_df):
    df = jobs_df.copy()
    df["Fill %"] = (df["Fill Rate"] * 100).round(1)
    df["Short Title"] = df["Job Title"].str[:35]
    df = df.sort_values("Fill %")
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df["Short Title"], y=df["# Openings"],
        name="Openings", marker_color="rgba(124,134,247,0.35)",
    ))
    fig.add_trace(go.Bar(
        x=df["Short Title"], y=df["Placements"],
        name="Placed", marker_color="#940000",
    ))
    fig.update_layout(
        barmode="overlay",
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=10, r=10, t=10, b=80),
        xaxis=dict(**DARK_AXIS, tickangle=-35, tickfont=dict(size=10)),
        yaxis=dict(**DARK_AXIS, title="Count"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
        height=320,
    )
    return fig


def make_fig_jobs_fillrate(jobs_df):
    df = jobs_df.copy()
    df["Fill %"] = (df["Fill Rate"] * 100).round(1)
    df["Short Title"] = df["Job Title"].str[:35]
    df = df.sort_values("Fill %")
    fig = px.bar(
        df, x="Short Title", y="Fill %",
        color="Fill %",
        color_continuous_scale=["#f77c8a", "#B2B5B8", "#B2B5B8"],
        range_color=[0, 100],
        text="Fill %",
    )
    fig.update_traces(texttemplate="%{text:.0f}%", textposition="outside")
    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=10, r=10, t=10, b=80),
        xaxis=dict(**DARK_AXIS, tickangle=-35, tickfont=dict(size=10)),
        yaxis=dict(**DARK_AXIS, title="Fill %", range=[0, 120]),
        coloraxis_showscale=False,
        height=300,
    )
    return fig


def make_fig_jobs_days(jobs_df):
    df = jobs_df.copy()
    df["Short Title"] = df["Job Title"].str[:35]
    df = df.sort_values("Days Opened", ascending=True)
    fig = px.bar(
        df, x="Days Opened", y="Short Title", orientation="h",
        color="Days Opened",
        color_continuous_scale=["#B2B5B8", "#B2B5B8", "#f77c8a"],
        text="Days Opened",
    )
    fig.update_traces(textposition="outside")
    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(**DARK_AXIS, title="Days Open"),
        yaxis=dict(**DARK_AXIS),
        coloraxis_showscale=False,
        height=320,
    )
    return fig


def make_fig_hires(hires_df):
    if hires_df.empty:
        return go.Figure()
    df = hires_df.copy()
    df["Week"] = df["Start Date"].dt.to_period("W").dt.start_time
    weekly = df.groupby("Week")["Name"].nunique().reset_index(name="New Hires")
    weekly["Week Label"] = weekly["Week"].dt.strftime("%b %d")
    _mv = weekly["New Hires"].max()
    max_val = int(_mv) if pd.notna(_mv) and _mv > 0 else 1
    fig = px.bar(weekly, x="Week Label", y="New Hires",
                 color_discrete_sequence=["#7D1F32"], text="New Hires")
    fig.update_traces(textposition="outside")
    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=10, r=10, t=30, b=10),
        xaxis=dict(**DARK_AXIS, title=""),
        yaxis=dict(**DARK_AXIS, title=dict(text="# Hires", font=dict(color="#131313")), range=[0, max_val * 1.2]),
        height=300,
    )
    return fig


def _grouped_bar_chart(curr_df, past_df, category_col, value_col, main_color, past_color="rgba(180,180,180,0.6)"):
    """Reusable grouped horizontal bar chart with optional prior period overlay."""
    curr = curr_df.groupby(category_col)[value_col].sum().reset_index(name="Current")
    if past_df is not None and not past_df.empty and category_col in past_df.columns:
        past = past_df.groupby(category_col)[value_col].sum().reset_index(name="Prior")
        merged = curr.merge(past, on=category_col, how="outer").fillna(0)
        merged = merged.sort_values("Current", ascending=True)
        max_val = int(max(merged["Current"].max(), merged["Prior"].max()))
        tick_step = max(1, round(max_val / 10))
        num_cats = len(merged)
        fig = go.Figure()
        fig.add_trace(go.Bar(
            y=merged[category_col], x=merged["Prior"],
            name="Prior Period", orientation="h",
            marker_color=past_color,
            text=merged["Prior"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
        ))
        fig.add_trace(go.Bar(
            y=merged[category_col], x=merged["Current"],
            name="Current Period", orientation="h",
            marker_color=main_color,
            text=merged["Current"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
        ))
        fig.update_layout(
            barmode="group",
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(color="#131313", size=12),
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(**DARK_AXIS, dtick=tick_step, range=[0, max_val * 1.3]),
            yaxis=dict(**DARK_AXIS),
            legend=dict(orientation="h", yanchor="bottom", y=1.02, font=dict(color="#131313")),
            height=max(280, num_cats * 52),
        )
    else:
        merged = curr.sort_values("Current", ascending=True)
        max_val = int(merged["Current"].max())
        tick_step = max(1, round(max_val / 10))
        num_cats = len(merged)
        fig = go.Figure(go.Bar(
            y=merged[category_col], x=merged["Current"],
            orientation="h", marker_color=main_color,
            text=merged["Current"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
        ))
        fig.update_layout(
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            font=dict(color="#131313", size=12),
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(**DARK_AXIS, dtick=tick_step, range=[0, max_val * 1.25]),
            yaxis=dict(**DARK_AXIS),
            showlegend=False,
            height=max(280, num_cats * 42),
        )
    return fig


def make_fig_retention_tiers(metrics, metrics_past=None, active_days=None):
    """Bar chart showing 7/30/60-day retention percentages."""
    tiers = metrics["retention_tiers"]
    total = metrics["total_new_starts"]
    if total == 0:
        return go.Figure()

    all_days   = [7, 30, 60]
    all_labels = ["7-Day", "30-Day", "60-Day"]
    all_colors = ["#7b0000", "#940000", "#c0392b"]
    if active_days is None:
        active_days = all_days

    days_keys = [d for d in all_days if d in active_days]
    labels    = [l for d, l in zip(all_days, all_labels) if d in active_days]
    colors    = [c for d, c in zip(all_days, all_colors) if d in active_days]

    # Filter out tiers with no data (pct=None)
    valid = [(d, l, c) for d, l, c in zip(days_keys, labels, colors) if tiers[d]["pct"] is not None]
    if not valid:
        return go.Figure()
    days_keys, labels, colors = zip(*valid)

    pcts      = [tiers[d]["pct"] for d in days_keys]
    retained  = [tiers[d]["retained"] for d in days_keys]
    eligible  = [tiers[d]["eligible"] for d in days_keys]
    early     = [tiers[d]["early_count"] for d in days_keys]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=labels, y=pcts, name="Current",
        marker_color=colors,
        text=[f"<b>{p:.0f}%</b><br><span style='font-size:13px'><b>{r} of {e}</b></span>" for p, r, e in zip(pcts, retained, eligible)],
        textposition="outside", textfont=dict(size=18, color="#131313"),
        hovertemplate="%{x}: %{y:.1f}% retained<br>%{customdata[0]} retained of %{customdata[1]} eligible<extra></extra>",
        customdata=list(zip(retained, eligible)),
    ))

    all_past_colors = ["rgba(123,0,0,0.35)", "rgba(148,0,0,0.35)", "rgba(192,57,43,0.35)"]
    past_colors = [c for d, c in zip(all_days, all_past_colors) if d in days_keys]
    if metrics_past is not None and metrics_past["total_new_starts"] > 0:
        past_tiers = metrics_past["retention_tiers"]
        past_pcts = [past_tiers[d]["pct"] if past_tiers[d]["pct"] is not None else 0 for d in days_keys]
        fig.add_trace(go.Bar(
            x=labels, y=past_pcts, name="Prior Period",
            marker_color=past_colors,
            text=[f"{p:.0f}%" if p else "" for p in past_pcts],
            textposition="outside", textfont=dict(size=11, color="#131313"),
        ))

    max_y = max(pcts) if pcts else 100
    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=14),
        margin=dict(l=40, r=20, t=50, b=70),
        yaxis=dict(range=[0, min(max_y * 1.35, 135)], title=dict(text="Retention %", font=dict(color="#131313")), showgrid=True,
                   gridcolor="#eee", ticksuffix="%", tickfont=dict(color="#131313", size=13)),
        xaxis=dict(title="", tickfont=dict(color="#131313", size=15)),
        barmode="group",
        showlegend=bool(metrics_past is not None and metrics_past["total_new_starts"] > 0),
        legend=dict(orientation="h", yanchor="top", y=-0.18, xanchor="center", x=0.5),
        height=380,
    )
    return fig


def make_fig_early_terms_by_reason(early_terms_df, title_suffix=""):
    """Horizontal bar chart of early terminations by end reason."""
    if early_terms_df.empty:
        return go.Figure()
    grouped = early_terms_df.groupby("End Reason")["Count"].sum().reset_index(name="Count")
    grouped = grouped.sort_values("Count", ascending=True)
    max_val = int(grouped["Count"].max()) if not grouped.empty else 1

    fig = go.Figure(go.Bar(
        y=grouped["End Reason"], x=grouped["Count"],
        orientation="h", marker_color="#c62828",
        text=grouped["Count"], textposition="outside",
    ))
    fig.update_layout(
        plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
        font=dict(color="#131313", size=12),
        margin=dict(l=10, r=10, t=10, b=10),
        xaxis=dict(**DARK_AXIS, range=[0, max_val * 1.3], dtick=max(1, round(max_val / 10))),
        yaxis=dict(**DARK_AXIS),
        showlegend=False,
        height=max(220, len(grouped) * 40),
    )
    return fig


def make_fig_rep_hires(hires_df, hires_past=None):
    if hires_df.empty or "Staffing Rep" not in hires_df.columns:
        return go.Figure()
    df = hires_df[hires_df["Staffing Rep"].str.strip() != "Unknown"].copy()
    if df.empty:
        df = hires_df.copy()
    df["_n"] = 1
    past_df = None
    if hires_past is not None and not hires_past.empty and "Staffing Rep" in hires_past.columns:
        past_df = hires_past[hires_past["Staffing Rep"].str.strip() != "Unknown"].copy()
        past_df["_n"] = 1
    return _grouped_bar_chart(df, past_df, "Staffing Rep", "_n", "#940000")


def make_fig_job_title_hires(hires_df, hires_past=None):
    if hires_df.empty or "Job Title" not in hires_df.columns:
        return go.Figure()
    df = hires_df[hires_df["Job Title"].str.strip() != "Unknown"].copy()
    if df.empty:
        df = hires_df.copy()
    df["_n"] = 1
    past_df = None
    if hires_past is not None and not hires_past.empty and "Job Title" in hires_past.columns:
        past_df = hires_past[hires_past["Job Title"].str.strip() != "Unknown"].copy()
        past_df["_n"] = 1
    return _grouped_bar_chart(df, past_df, "Job Title", "_n", "#7D1F32")


def strip_html(val):
    import re, html
    if pd.isna(val): return ""
    return html.unescape(re.sub(r"<[^>]+>", " ", str(val))).strip()

def format_terms_display(df):
    if df.empty:
        return df
    d = df.copy()
    cols = ["Name", "End Reason", "Comments", "Start Date", "End Date"]
    cols = [c for c in cols if c in d.columns]
    for col in ["Start Date", "End Date"]:
        if col in d.columns:
            d[col] = pd.to_datetime(d[col], errors="coerce").dt.strftime("%b %d, %Y").fillna("")
    if "Comments" in d.columns:
        d["Comments"] = d["Comments"].apply(strip_html)
    return d[cols]


PDF_BG     = "white"
PDF_GRID   = "#e8e8e8"
PDF_FONT   = "#131313"
PDF_ACCENT  = "#940000"   # KP crimson
PDF_ACCENT2 = "#7D1F32"   # KP dark maroon

def _pdf_layout(extra=None):
    base = dict(
        plot_bgcolor=PDF_BG, paper_bgcolor=PDF_BG,
        font=dict(color=PDF_FONT, family="Helvetica, Arial, sans-serif", size=11),
        margin=dict(l=160, r=30, t=20, b=50),
        height=300,
    )
    if extra:
        base.update(extra)
    return base

def make_pdf_fig_headcount(hc_df):
    hc = hc_df.copy()
    hc["Week"] = hc["Week Ending"].dt.strftime("%b %d")
    max_hc = float(hc["Headcount"].max()) if not hc.empty else 1.0
    min_hc = float(hc["Headcount"].min()) if not hc.empty else 0.0
    has_hours = "Hours" in hc.columns and float(hc["Hours"].sum()) > 0

    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=hc["Week"], y=hc["Headcount"],
        name="Headcount",
        mode="lines+markers",
        line=dict(color=PDF_ACCENT, width=2.5),
        marker=dict(size=8, color=PDF_ACCENT),
        fillcolor="rgba(148,0,0,0.08)",
        yaxis="y1",
    ))
    num_pts = len(hc)
    for idx, (_, row) in enumerate(hc.iterrows()):
        # Push edge labels inward so they don't get clipped
        xs = 20 if idx == 0 else (-20 if idx == num_pts - 1 else 0)
        fig.add_annotation(
            x=row["Week"], y=float(row["Headcount"]),
            text=str(int(row["Headcount"])),
            showarrow=False, yshift=14, xshift=xs,
            font=dict(size=11, color=PDF_ACCENT),
            xref="x", yref="y",
        )

    layout_extra = {
        "margin": dict(l=60, r=80 if has_hours else 70, t=40, b=50),
        "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True),
        "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                      title=dict(text="Headcount", font=dict(color=PDF_ACCENT)),
                      range=[max(0, min_hc * 0.85), max_hc * 1.15]),
        "height": 260,
        "showlegend": True if has_hours else False,
        "legend": dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=9)),
    }

    if has_hours:
        max_hrs = float(hc["Hours"].max())
        fig.add_trace(go.Scatter(
            x=hc["Week"], y=hc["Hours"],
            name="Hours",
            mode="lines+markers",
            line=dict(color="#1a6ea8", width=2, dash="dot"),
            marker=dict(size=6, color="#1a6ea8"),
            yaxis="y2",
        ))
        for idx2, (_, row) in enumerate(hc.iterrows()):
            xs2 = 20 if idx2 == 0 else (-20 if idx2 == num_pts - 1 else 0)
            fig.add_annotation(
                x=row["Week"], y=float(row["Hours"]),
                text=f"{int(row['Hours']):,}",
                showarrow=False, yshift=-16, xshift=xs2,
                font=dict(size=10, color="#1a6ea8"),
                xref="x", yref="y2",
            )
        layout_extra["yaxis2"] = dict(
            title=dict(text="Hours", font=dict(color="#1a6ea8")),
            overlaying="y", side="right",
            showgrid=False,
            tickfont=dict(color="#1a6ea8"),
            range=[0, max_hrs * 1.25],
        )

    fig.update_layout(**_pdf_layout(layout_extra))
    return fig

def make_pdf_fig_donut(m):
    types = ["Involuntary", "Voluntary", "Layoff / Assignment Complete"]
    counts = [m["inval_count"], m["vol_count"], m["layoff_count"]]
    clrs = ["#940000", "#7D1F32", "#B2B5B8"]
    fig = px.pie(
        pd.DataFrame({"Type": types, "Count": counts}),
        names="Type", values="Count",
        color_discrete_sequence=clrs, hole=0.5,
    )
    fig.update_traces(textinfo="percent+value", textfont_size=13)
    fig.update_layout(
        plot_bgcolor=PDF_BG, paper_bgcolor=PDF_BG,
        font=dict(color=PDF_FONT, family="Helvetica, Arial, sans-serif", size=11),
        margin=dict(l=10, r=10, t=10, b=80),
        height=380,
        legend=dict(orientation="h", yanchor="bottom", y=-0.2,
                    xanchor="center", x=0.5, font=dict(size=11)),
    )
    return fig

def make_pdf_fig_reasons(df, bar_color):
    if df.empty:
        return go.Figure()
    grouped = df.groupby("End Reason")["Count"].sum().reset_index().sort_values("Count")
    # Wrap long labels
    grouped["Label"] = grouped["End Reason"].str.replace("/", "/<br>")
    fig = go.Figure(go.Bar(
        x=grouped["Count"], y=grouped["Label"],
        orientation="h",
        marker_color=bar_color,
        text=grouped["Count"], textposition="outside",
        textfont=dict(size=11),
    ))
    max_count = grouped["Count"].max() if not grouped.empty else 1
    fig.update_layout(**_pdf_layout({
        "margin": dict(l=200, r=60, t=20, b=40),
        "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                      range=[0, max_count * 1.25], dtick=1),
        "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                      tickfont=dict(size=10)),
        "showlegend": False,
    }))
    return fig

def make_pdf_fig_hires(hires_df):
    if hires_df.empty:
        return go.Figure()
    df = hires_df.copy()
    df["Week"] = df["Start Date"].dt.to_period("W").dt.start_time
    weekly = df.groupby("Week")["Name"].nunique().reset_index(name="New Hires")
    weekly["Label"] = weekly["Week"].dt.strftime("%b %d")
    fig = go.Figure(go.Bar(
        x=weekly["Label"], y=weekly["New Hires"],
        marker_color=PDF_ACCENT2,
        text=weekly["New Hires"], textposition="outside",
        textfont=dict(size=11),
    ))
    max_val = weekly["New Hires"].max() if not weekly.empty else 1
    fig.update_layout(**_pdf_layout({
        "margin": dict(l=50, r=30, t=20, b=50),
        "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True),
        "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                      title="Hires", range=[0, max_val * 1.2]),
    }))
    return fig


def generate_excel(data, metrics, data_past=None, metrics_past=None):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    # ── Color palette matching PDF ────────────────────────────────────────────
    C_RED    = "940000"
    C_MAROON = "7D1F32"
    C_LIGHT  = "fdf0f0"
    C_STRIPE = "faf7f7"
    C_GRID   = "e8d8d8"
    C_WHITE  = "FFFFFF"
    C_GREY   = "888888"
    C_GREEN  = "2e7d32"
    C_DKRED  = "c62828"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border(color=C_GRID, top=False, bottom=False, left=False, right=False):
        s = Side(style="thin", color=color)
        n = None
        return Border(top=s if top else n, bottom=s if bottom else n,
                      left=s if left else n, right=s if right else n)

    def thick_top(color=C_RED):
        return Border(top=Side(style="medium", color=color))

    def all_border():
        s = Side(style="thin", color=C_GRID)
        return Border(top=s, bottom=s, left=s, right=s)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: Report Summary  (mirrors PDF layout)
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Report Summary"

    # Column widths
    for col, w in [(1,28),(2,18),(3,18),(4,18)]:
        ws.column_dimensions[get_column_letter(col)].width = w

    tiers = metrics["retention_tiers"]
    m_past = metrics_past or {}

    def _delta_str(curr, prev, is_pct=False):
        if prev is None or prev == 0:
            return ""
        diff = curr - prev
        sign = "▲" if diff > 0 else ("▼" if diff < 0 else "—")
        if is_pct:
            return f"{sign} {abs(diff):.1f}pp vs prior"
        return f"{sign} {abs(diff):g} vs prior"

    def _delta_color(curr, prev, good="up"):
        if prev is None:
            return None
        diff = curr - prev
        if abs(diff) < 0.05:
            return C_GREY
        if good == "up":
            return C_GREEN if diff > 0 else C_DKRED
        return C_DKRED if diff > 0 else C_GREEN

    row = 1

    # ── Header ────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "KP Staffing — Activity Report")
    c.font = Font(bold=True, size=16, color=C_WHITE)
    c.fill = fill(C_RED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    ws.cell(row, 1, data.get("company", "")).font = Font(bold=True, size=12)
    ws.cell(row, 1).alignment = Alignment(indent=1)
    ws.merge_cells(f"C{row}:D{row}")
    ws.cell(row, 3, data.get("date_range", "")).font = Font(size=10, color=C_GREY)
    ws.cell(row, 3).alignment = Alignment(horizontal="right")
    ws.row_dimensions[row].height = 20
    row += 2

    def section_header(label):
        nonlocal row
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row, 1, f"  {label}")
        c.font = Font(bold=True, size=10, color=C_WHITE)
        c.fill = fill(C_MAROON)
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    def kpi_card(col, value, label, subtitle="", delta_val=None, delta_prev=None, good="up", is_pct=False):
        """Write a KPI card starting at (row, col). Uses 3 rows."""
        nonlocal row
        r = row
        # Value row
        c = ws.cell(r, col, str(value))
        c.font = Font(bold=True, size=18, color=C_RED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thick_top(C_RED)
        c.fill = fill(C_WHITE)
        ws.row_dimensions[r].height = 28
        # Label row
        c2 = ws.cell(r+1, col, label)
        c2.font = Font(bold=True, size=10)
        c2.alignment = Alignment(horizontal="center")
        c2.fill = fill(C_LIGHT)
        ws.row_dimensions[r+1].height = 16
        # Subtitle / delta row
        sub_text = subtitle
        sub_color = C_GREY
        if delta_val is not None and delta_prev is not None:
            d_str = _delta_str(delta_val, delta_prev, is_pct)
            d_col = _delta_color(delta_val, delta_prev, good)
            if d_str:
                sub_text = d_str
                sub_color = d_col or C_GREY
        c3 = ws.cell(r+2, col, sub_text)
        c3.font = Font(size=8, color=sub_color)
        c3.alignment = Alignment(horizontal="center")
        c3.fill = fill(C_LIGHT)
        ws.row_dimensions[r+2].height = 14

    def card_row(cards):
        """cards = list of (value, label, subtitle, delta_val, delta_prev, good, is_pct)"""
        nonlocal row
        for i, card in enumerate(cards):
            kpi_card(i+1, *card)
        row += 3

    # ── Row 1: Workforce ──────────────────────────────────────────────────────
    section_header("WORKFORCE")
    _ph = metrics["starting_headcount"]
    _ph_p = m_past.get("starting_headcount")
    _ps = metrics["total_new_starts"]
    _ps_p = m_past.get("total_new_starts")
    _hrs = f"{metrics['total_hours']:,.0f}"
    _hrs_p = m_past.get("total_hours")
    cards1 = [
        (_ph, "Starting Headcount", "at period start", _ph, _ph_p, "up", False),
        (_ps, "Total Starts", "new placements", _ps, _ps_p, "up", False),
        (_hrs, "Total Hours", "billed this period", metrics["total_hours"], _hrs_p, "up", False),
    ]
    card_row(cards1)
    row += 1

    # ── Row 2: Turnover ───────────────────────────────────────────────────────
    section_header("TURNOVER")
    _ic = metrics["inval_count"]; _ic_p = m_past.get("inval_count")
    _vc = metrics["vol_count"];   _vc_p = m_past.get("vol_count")
    _tp = metrics["total_turnover_pct"]; _tp_p = m_past.get("total_turnover_pct")
    cards2 = [
        (f"{metrics['inval_pct']:.1f}%", "Involuntary Terms",
         f"{_ic} employees", _ic, _ic_p, "down", False),
        (f"{metrics['vol_pct']:.1f}%", "Voluntary Terms",
         f"{_vc} employees", _vc, _vc_p, "down", False),
        (f"{_tp:.1f}%", "Total Turnover",
         "excl. layoffs and conversions", _tp, _tp_p, "down", True),
    ]
    card_row(cards2)
    row += 1

    # ── Row 3: Retention ──────────────────────────────────────────────────────
    section_header("RETENTION")
    ret_cards = []
    for days in (7, 30, 60):
        t = tiers[days]
        if t["pct"] is None:
            ret_cards.append(("N/A", f"{days}-Day Retention", "Insufficient data", None, None, "up", True))
        else:
            _rp = t["pct"]; _rp_p = (m_past.get("retention_tiers") or {}).get(days, {}).get("pct")
            ret_cards.append((f"{_rp:.0f}%", f"{days}-Day Retention",
                              f"{t['retained']} of {t['eligible']} eligible",
                              _rp, _rp_p, "up", True))
    card_row(ret_cards)
    row += 1

    # ── Row 4: Exclusions ─────────────────────────────────────────────────────
    section_header("EXCLUSIONS & CONVERSIONS")
    excl_cards = [
        (metrics["layoff_count"], "Layoffs / Assign. Complete",
         "excluded from turnover",
         metrics["layoff_count"], m_past.get("layoff_count"), "neutral", False),
    ]
    if not data["converted"].empty:
        excl_cards.append((len(data["converted"]), "Converted Employees",
                           "hired by client this period", None, None, "up", False))
    card_row(excl_cards)
    row += 2

    # ── Termination tables ────────────────────────────────────────────────────
    def write_table(df, title, start_row):
        if df.empty:
            return start_row
        cols = [c for c in ["Name","End Reason","Comments","Start Date","End Date"] if c in df.columns]
        r = start_row
        ws.merge_cells(f"A{r}:D{r}")
        c = ws.cell(r, 1, title)
        c.font = Font(bold=True, size=11, color=C_WHITE)
        c.fill = fill(C_RED)
        c.alignment = Alignment(indent=1, vertical="center")
        ws.row_dimensions[r].height = 18
        r += 1
        for ci, col in enumerate(cols, 1):
            c = ws.cell(r, ci, col)
            c.font = Font(bold=True, size=9, color="131313")
            c.fill = fill("e8edf2")
            c.border = all_border()
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 14
        r += 1
        for idx, (_, row_data) in enumerate(df.iterrows()):
            bg = C_WHITE if idx % 2 == 0 else C_STRIPE
            for ci, col in enumerate(cols, 1):
                val = row_data.get(col, "")
                if hasattr(val, "strftime"):
                    val = val.strftime("%m/%d/%Y")
                c = ws.cell(r, ci, str(val) if pd.notna(val) else "")
                c.font = Font(size=9)
                c.fill = fill(bg)
                c.border = all_border()
            r += 1
        return r + 1

    row = write_table(metrics["involuntary"], "Involuntary Terminations", row)
    row = write_table(metrics["voluntary"],   "Voluntary Terminations",   row)
    row = write_table(metrics["layoffs"],     "Layoffs / Assign. Complete", row)

    # Conversions table on Summary sheet
    if not data["converted"].empty:
        _conv_xl = data["converted"].copy()
        _conv_xl_cols = [c for c in ["Name", "Job Title", "Staffing Rep", "Converted Date", "Start Date"] if c in _conv_xl.columns]
        for _dc in ["Converted Date", "Start Date"]:
            if _dc in _conv_xl.columns:
                _conv_xl[_dc] = pd.to_datetime(_conv_xl[_dc], errors="coerce").dt.strftime("%m/%d/%Y")
        row = write_table(_conv_xl[_conv_xl_cols], "Conversions to Permanent", row)

    # Early term detail
    for days in (60, 30, 7):
        t = tiers[days]
        if t["pct"] is not None and not t["early_terms"].empty:
            row = write_table(t["early_terms"], f"Early Exits — Within {days} Days", row)
            break

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: New Hires
    # ═══════════════════════════════════════════════════════════════════════════
    if not data["hires"].empty:
        ws2 = wb.create_sheet("New Hires")
        _write_raw_sheet(ws2, data["hires"], "New Hires", C_RED, C_LIGHT, C_STRIPE, C_GRID)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: Terminations
    # ═══════════════════════════════════════════════════════════════════════════
    if not data["terms"].empty:
        ws3 = wb.create_sheet("Terminations")
        _write_raw_sheet(ws3, data["terms"], "Terminations", C_RED, C_LIGHT, C_STRIPE, C_GRID)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4: Headcount by Week
    # ═══════════════════════════════════════════════════════════════════════════
    if not data["headcount"].empty:
        ws4 = wb.create_sheet("Headcount by Week")
        _write_raw_sheet(ws4, data["headcount"], "Headcount by Week", C_RED, C_LIGHT, C_STRIPE, C_GRID)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 5: Conversions
    # ═══════════════════════════════════════════════════════════════════════════
    if not data["converted"].empty:
        ws5 = wb.create_sheet("Conversions")
        _conv_s = data["converted"].copy()
        _conv_s_cols = [c for c in ["Name", "Job Title", "Staffing Rep", "Converted Date", "Start Date"] if c in _conv_s.columns]
        for _dc in ["Converted Date", "Start Date"]:
            if _dc in _conv_s.columns:
                _conv_s[_dc] = pd.to_datetime(_conv_s[_dc], errors="coerce").dt.strftime("%m/%d/%Y")
        _write_raw_sheet(ws5, _conv_s[_conv_s_cols], "Conversions to Permanent", C_RED, C_LIGHT, C_STRIPE, C_GRID)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 6: Jobs
    # ═══════════════════════════════════════════════════════════════════════════
    if not data.get("jobs", pd.DataFrame()).empty:
        ws6 = wb.create_sheet("Jobs")
        _write_raw_sheet(ws6, data["jobs"], "Jobs Overview", C_RED, C_LIGHT, C_STRIPE, C_GRID)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_raw_sheet(ws, df, title, C_RED, C_LIGHT, C_STRIPE, C_GRID):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def fill(h): return PatternFill("solid", fgColor=h)
    def all_border():
        s = Side(style="thin", color=C_GRID)
        return Border(top=s, bottom=s, left=s, right=s)

    cols = list(df.columns)
    # Header row
    ws.cell(1, 1, title).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(1, 1).fill = fill(C_RED)
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws.row_dimensions[1].height = 20
    # Column headers
    for ci, col in enumerate(cols, 1):
        c = ws.cell(2, ci, col)
        c.font = Font(bold=True, size=9)
        c.fill = fill("e8edf2")
        c.border = all_border()
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(str(col)) + 2)
    ws.row_dimensions[2].height = 14
    # Data rows
    for ri, (_, row) in enumerate(df.iterrows()):
        bg = "FFFFFF" if ri % 2 == 0 else C_STRIPE
        for ci, col in enumerate(cols, 1):
            val = row[col]
            if hasattr(val, "strftime"):
                val = val.strftime("%m/%d/%Y")
            c = ws.cell(ri+3, ci, str(val) if pd.notna(val) else "")
            c.font = Font(size=9)
            c.fill = fill(bg)
            c.border = all_border()


def generate_pdf(data, metrics, data_past=None, metrics_past=None, display_opts=None):
    import re, html as html_lib
    # Default all display options to True if not provided
    opts = display_opts or {}
    _show_hours     = opts.get("show_hours",     True)
    _show_converted = opts.get("show_converted", True)
    _show_hc_chart  = opts.get("show_hc_chart",  True)
    _show_donut     = opts.get("show_donut",      True)
    _show_reasons   = opts.get("show_reasons",    True)
    _show_hires     = opts.get("show_hires",      True)
    _show_rep       = opts.get("show_rep",        True)
    _show_title     = opts.get("show_title",      True)
    _show_retention    = opts.get("show_retention",    True)
    _show_retention_7  = opts.get("show_retention_7",  True)
    _show_retention_30 = opts.get("show_retention_30", True)
    _show_retention_60 = opts.get("show_retention_60", True)
    _pdf_active_days   = [d for d, s in [(7, _show_retention_7), (30, _show_retention_30), (60, _show_retention_60)] if s]
    _show_jobs      = opts.get("show_jobs",       True)
    _show_tables    = opts.get("show_tables",     True)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.6*inch, rightMargin=0.6*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    # ── Styles ──────────────────────────────────────────────────────────────
    NAVY  = colors.HexColor("#940000")   # KP crimson
    SLATE = colors.HexColor("#7D1F32")   # KP dark maroon
    LIGHT = colors.HexColor("#fdf0f0")   # KP tinted light
    STRIPE= colors.HexColor("#faf7f7")
    GRID  = colors.HexColor("#e8d8d8")

    title_s = ParagraphStyle("T",  fontSize=14, textColor=NAVY,
                              fontName="Helvetica-Bold", spaceAfter=3)
    sub_s   = ParagraphStyle("S",  fontSize=9,  textColor=colors.HexColor("#666677"),
                              spaceAfter=10)
    h2_s    = ParagraphStyle("H2", fontSize=10, textColor=NAVY,
                              fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=5)
    cell_s  = ParagraphStyle("C",  fontSize=7.5, leading=10)
    th_s    = ParagraphStyle("TH", fontSize=8,  fontName="Helvetica-Bold",
                              textColor=colors.white, leading=10)

    def safe_cell(val):
        if pd.isna(val): return ""
        if hasattr(val, "strftime"): return val.strftime("%b %d, %Y")
        text = re.sub(r"<[^>]+>", " ", str(val))
        return html_lib.escape(html_lib.unescape(text).strip())

    story = []

    # ── Header with logo ─────────────────────────────────────────────────────
    logo_path = "kp_logo.png"
    try:
        logo_img = RLImage(logo_path, width=1.6*inch, height=0.4*inch)
        header_data = [[logo_img,
                        Paragraph(f"{data['company']}  —  Activity Report", title_s)]]
        header_t = Table(header_data, colWidths=[1.8*inch, 8.2*inch])
        header_t.setStyle(TableStyle([
            ("VALIGN",  (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING",  (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
            ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ]))
        story.append(header_t)
    except Exception:
        story.append(Paragraph(f"{data['company']}  —  Activity Report", title_s))
    _cur_range = display_opts.get("period_label", data["date_range"]) if display_opts else data["date_range"]
    period_line = f"Period: {_cur_range}"
    if data_past is not None:
        _past_range = display_opts.get("past_period_label", data_past["date_range"]) if display_opts else data_past["date_range"]
        period_line += f"   |   Prior Period: {_past_range}"
    story.append(Paragraph(period_line, sub_s))
    story.append(HRFlowable(width="100%", thickness=1.5, color=NAVY))
    story.append(Spacer(1, 10))

    # ── KPI Cards (matching dashboard style) ───────────────────────────────
    story.append(Paragraph("Key Metrics", h2_s))

    def _pdf_delta(diff, is_pct=False, good="down"):
        """Returns (text, paragraph_style) tuple for a delta value."""
        if abs(diff) < (0.05 if is_pct else 0.5):
            return "No change vs prior", card_delta_neu_s
        sign  = "+" if diff > 0 else ""
        arrow = "+" if diff > 0 else "-"
        val   = f"{sign}{diff:.1f}pp" if is_pct else f"{sign}{int(round(diff))}"
        text  = f"{arrow} {val} vs prior"
        if good == "neutral":
            style = card_delta_neu_s
        elif good == "up":
            style = card_delta_good_s if diff > 0 else card_delta_bad_s
        else:  # good == "down"
            style = card_delta_bad_s if diff > 0 else card_delta_good_s
        return text, style

    curr_hc  = metrics["starting_headcount"] or metrics["adj_placements"]
    total_hrs_str = f"{int(metrics['total_hours']):,}" if metrics["total_hours"] > 0 else "0"
    converted_count = len(data["converted"]) if not data["converted"].empty else 0
    converted_count = converted_count if _show_converted else 0

    card_val_s   = ParagraphStyle("cv", fontSize=14, fontName="Helvetica-Bold",
                                   textColor=NAVY, alignment=TA_CENTER, spaceAfter=4, leading=16)
    card_label_s = ParagraphStyle("cl", fontSize=6, fontName="Helvetica-Bold",
                                   textColor=colors.black, alignment=TA_CENTER,
                                   spaceAfter=3, leading=8)
    card_sub_s   = ParagraphStyle("cs", fontSize=6, textColor=colors.black,
                                   alignment=TA_CENTER, leading=8)
    card_pct_s   = ParagraphStyle("cp", fontSize=11, fontName="Helvetica-Bold",
                                   textColor=SLATE, alignment=TA_CENTER, spaceBefore=2, spaceAfter=1, leading=13)
    card_delta_good_s = ParagraphStyle("cdg", fontSize=6, fontName="Helvetica-Bold",
                                       textColor=colors.HexColor("#2e7d32"), alignment=TA_CENTER, leading=8)
    card_delta_bad_s  = ParagraphStyle("cdb", fontSize=6, fontName="Helvetica-Bold",
                                       textColor=colors.HexColor("#c62828"), alignment=TA_CENTER, leading=8)
    card_delta_neu_s  = ParagraphStyle("cdn", fontSize=6, fontName="Helvetica-Bold",
                                       textColor=colors.HexColor("#888888"), alignment=TA_CENTER, leading=8)
    # Keep backward-compat alias
    card_delta_s = card_delta_neu_s

    def _make_card_cell(value, label, sub, pct=None, delta_txt="", delta_style=None):
        """Build a list of flowables for one KPI card cell."""
        parts = [
            Paragraph(str(value), card_val_s),
            Paragraph(label.upper(), card_label_s),
        ]
        if pct is not None:
            parts.append(HRFlowable(width="50%", thickness=0.5, color=colors.HexColor("#e8e8e8"),
                                     spaceBefore=3, spaceAfter=3))
            parts.append(Paragraph(str(pct), card_pct_s))
        if delta_txt:
            parts.append(Paragraph(delta_txt, delta_style or card_delta_neu_s))
        parts.append(Paragraph(sub, card_sub_s))
        return parts

    kpi_style_cmds = [
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("BOX",           (0, 0), (-1, -1), 0.5, GRID),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, GRID),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("BACKGROUND",    (0, 0), (-1, -1), colors.white),
        ("LINEABOVE",     (0, 0), (-1, 0),  3, NAVY),
    ]

    def _render_kpi_row(cells):
        if not cells:
            return
        n = len(cells)
        w = 9.8 * inch / n
        t = Table([cells], colWidths=[w] * n)
        t.setStyle(TableStyle(kpi_style_cmds))
        story.append(t)
        story.append(Spacer(1, 4))

    def _add_delta(cell, diff, is_pct, good):
        if metrics_past is not None:
            dt_text, dt_style = _pdf_delta(diff, is_pct=is_pct, good=good)
            cell.insert(-1, Paragraph(dt_text, dt_style))

    p = metrics_past or {}
    past_hc = (p.get("starting_headcount") or p.get("adj_placements", 0)) if metrics_past else 0

    # ── Row 1: Workforce ──────────────────────────────────────────────────────
    row1 = []
    hc_cell = _make_card_cell(curr_hc, "Starting Headcount", "headcount at period start")
    _add_delta(hc_cell, curr_hc - past_hc, False, "up")
    row1.append(hc_cell)

    starts_cell = _make_card_cell(data["new_placements_raw"], "Total Starts", "new placements this period")
    if metrics_past and data_past:
        _add_delta(starts_cell, data["new_placements_raw"] - data_past["new_placements_raw"], False, "up")
    row1.append(starts_cell)

    if _show_hours:
        hrs_cell = _make_card_cell(total_hrs_str, "Total Hours", "billed this period")
        if metrics_past:
            _add_delta(hrs_cell, metrics["total_hours"] - p["total_hours"], False, "up")
        row1.append(hrs_cell)

    _render_kpi_row(row1)

    # ── Row 2: Turnover ───────────────────────────────────────────────────────
    row2 = []
    inval_cell = _make_card_cell(metrics["inval_count"], "Involuntary Terms", "of total active employees",
                                 pct=f"{metrics['inval_pct']:.1f}%")
    if metrics_past:
        _add_delta(inval_cell, metrics["inval_count"] - p["inval_count"], False, "down")
    row2.append(inval_cell)

    vol_cell = _make_card_cell(metrics["vol_count"], "Voluntary Terms", "of total active employees",
                               pct=f"{metrics['vol_pct']:.1f}%")
    if metrics_past:
        _add_delta(vol_cell, metrics["vol_count"] - p["vol_count"], False, "down")
    row2.append(vol_cell)

    to_cell = _make_card_cell(f"{metrics['total_turnover_pct']:.1f}%", "Total Turnover",
                              "excl. layoffs and conversions",
                              pct=f"{metrics['inval_count'] + metrics['vol_count']} total terms")
    if metrics_past:
        _add_delta(to_cell, metrics["total_turnover_pct"] - p["total_turnover_pct"], True, "down")
    row2.append(to_cell)

    _render_kpi_row(row2)

    # ── Row 3: Retention ──────────────────────────────────────────────────────
    row3 = []
    tiers = metrics["retention_tiers"]
    for _days in _pdf_active_days if _pdf_active_days else []:
        _tier = tiers[_days]
        if _tier["pct"] is None:
            ret_cell = _make_card_cell("N/A", f"{_days}-Day Retention", "Insufficient data")
        else:
            ret_cell = _make_card_cell(
                f"{_tier['pct']:.0f}%", f"{_days}-Day Retention",
                f"{_tier['retained']} of {_tier['eligible']} eligible"
            )
            if metrics_past and metrics_past["total_new_starts"] > 0:
                _past_tier = metrics_past["retention_tiers"][_days]
                if _past_tier["pct"] is not None:
                    _add_delta(ret_cell, _tier["pct"] - _past_tier["pct"], True, "up")
        row3.append(ret_cell)
    _render_kpi_row(row3)

    # ── Row 4: Exclusions & Conversions ──────────────────────────────────────
    row4 = []
    layoff_cell = _make_card_cell(metrics["layoff_count"], "Layoffs / Assign. Complete", "excluded from turnover")
    if metrics_past:
        _add_delta(layoff_cell, metrics["layoff_count"] - p["layoff_count"], False, "neutral")
    row4.append(layoff_cell)

    if converted_count > 0:
        row4.append(_make_card_cell(converted_count, "Converted Employees", "hired by client this period"))
    _render_kpi_row(row4)

    story.append(Spacer(1, 10))

    # ── Charts ────────────────────────────────────────────────────────────────
    def add_chart(fig, label, w=9.8*inch, h=2.5*inch, center=False):
        try:
            import plotly.io as pio
            px_w = int(w / inch * 110)
            px_h = int(h / inch * 110)
            img_bytes = pio.to_image(fig, format="png", width=px_w, height=px_h, scale=2, engine="kaleido")
            story.append(Paragraph(label, h2_s))
            img = RLImage(io.BytesIO(img_bytes), width=w, height=h)
            if center:
                img.hAlign = "CENTER"
            story.append(img)
            story.append(Spacer(1, 8))
        except Exception as e:
            story.append(Paragraph(label, h2_s))
            story.append(Paragraph(f"[Chart unavailable: {e}]", styles["Normal"]))
            story.append(Spacer(1, 8))

    if _show_hc_chart and not data["headcount"].empty:
        story.append(PageBreak())
        add_chart(make_pdf_fig_headcount(data["headcount"]), "Headcount by Week", h=2.4*inch)
    if _show_donut:
        add_chart(make_pdf_fig_donut(metrics), "Terminations by Type", w=5.5*inch, h=3.2*inch, center=True)

    # Reason charts — use grouped bars when past data available
    _pi = metrics_past["involuntary"] if metrics_past else None
    _pv = metrics_past["voluntary"]   if metrics_past else None
    inval_label = "Involuntary Termination Reasons" + (" (Current vs Prior)" if _pi is not None else "")
    vol_label   = "Voluntary Termination Reasons"   + (" (Current vs Prior)" if _pv is not None else "")

    def _pdf_reasons_grouped(df, bar_color, df_past=None):
        if df.empty:
            return go.Figure()
        curr = df[df["End Reason"].str.strip() != ""].groupby("End Reason")["Count"].sum().reset_index(name="Current")
        if df_past is not None and not df_past.empty:
            past_f = df_past[df_past["End Reason"].str.strip() != ""]
            past   = past_f.groupby("End Reason")["Count"].sum().reset_index(name="Prior")
            merged = curr.merge(past, on="End Reason", how="outer").fillna(0).sort_values("Current", ascending=True)
            max_c  = int(max(merged["Current"].max(), merged["Prior"].max()))
            fig = go.Figure()
            fig.add_trace(go.Bar(y=merged["End Reason"], x=merged["Prior"],
                                 name="Prior", orientation="h",
                                 marker_color="rgba(180,180,180,0.7)",
                                 text=merged["Prior"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
                                 textfont=dict(size=9)))
            fig.add_trace(go.Bar(y=merged["End Reason"], x=merged["Current"],
                                 name="Current", orientation="h",
                                 marker_color=bar_color,
                                 text=merged["Current"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
                                 textfont=dict(size=9)))
            fig.update_layout(**_pdf_layout({
                "barmode": "group",
                "margin": dict(l=200, r=60, t=20, b=40),
                "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              range=[0, max_c * 1.3], dtick=max(1, round(max_c/10))),
                "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              tickfont=dict(size=10)),
                "legend": dict(orientation="h", yanchor="bottom", y=1.02,
                               font=dict(size=9)),
            }))
        else:
            grouped = curr.sort_values("Current", ascending=True)
            grouped["Label"] = grouped["End Reason"].str.replace("/", "/<br>")
            max_c = int(grouped["Current"].max()) if not grouped.empty else 1
            fig = go.Figure(go.Bar(
                x=grouped["Current"], y=grouped["Label"], orientation="h",
                marker_color=bar_color,
                text=grouped["Current"], textposition="outside", textfont=dict(size=11),
            ))
            fig.update_layout(**_pdf_layout({
                "margin": dict(l=200, r=60, t=20, b=40),
                "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              range=[0, max_c * 1.25], dtick=1),
                "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              tickfont=dict(size=10)),
                "showlegend": False,
            }))
        return fig

    if _show_reasons:
        add_chart(_pdf_reasons_grouped(metrics["involuntary"], "#940000", _pi), inval_label)
        add_chart(_pdf_reasons_grouped(metrics["voluntary"],   "#7D1F32", _pv), vol_label)

    if _show_hires and not data["hires"].empty:
        add_chart(make_pdf_fig_hires(data["hires"]), "New Hires by Start Week")

    # ── Staffing Rep + Job Title charts ───────────────────────────────────────
    hires_past_pdf = data_past["hires"] if data_past is not None else None

    def _pdf_grouped_bar(curr_df, past_df, category_col, bar_color):
        if curr_df.empty or category_col not in curr_df.columns:
            return None
        df = curr_df[curr_df[category_col].str.strip() != "Unknown"].copy()
        if df.empty:
            df = curr_df.copy()
        df["_n"] = 1
        curr = df.groupby(category_col)["_n"].sum().reset_index(name="Current")
        has_past = (past_df is not None and not past_df.empty and category_col in past_df.columns)
        if has_past:
            pdf2 = past_df[past_df[category_col].str.strip() != "Unknown"].copy()
            pdf2["_n"] = 1
            past = pdf2.groupby(category_col)["_n"].sum().reset_index(name="Prior")
            merged = curr.merge(past, on=category_col, how="outer").fillna(0).sort_values("Current", ascending=True)
            max_v = int(max(merged["Current"].max(), merged["Prior"].max()))
            fig = go.Figure()
            fig.add_trace(go.Bar(y=merged[category_col], x=merged["Prior"],
                                 name="Prior", orientation="h",
                                 marker_color="rgba(180,180,180,0.7)",
                                 text=merged["Prior"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
                                 textfont=dict(size=9)))
            fig.add_trace(go.Bar(y=merged[category_col], x=merged["Current"],
                                 name="Current", orientation="h",
                                 marker_color=bar_color,
                                 text=merged["Current"].apply(lambda v: int(v) if v > 0 else ""), textposition="outside",
                                 textfont=dict(size=9)))
            fig.update_layout(**_pdf_layout({
                "barmode": "group",
                "margin": dict(l=200, r=60, t=20, b=40),
                "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              range=[0, max_v * 1.3], dtick=max(1, round(max_v/10))),
                "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              tickfont=dict(size=10)),
                "legend": dict(orientation="h", yanchor="bottom", y=1.02, font=dict(size=9)),
            }))
        else:
            merged = curr.sort_values("Current", ascending=True)
            max_v = int(merged["Current"].max()) if not merged.empty else 1
            fig = go.Figure(go.Bar(
                x=merged["Current"], y=merged[category_col], orientation="h",
                marker_color=bar_color,
                text=merged["Current"], textposition="outside", textfont=dict(size=11),
            ))
            fig.update_layout(**_pdf_layout({
                "margin": dict(l=200, r=60, t=20, b=40),
                "xaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              range=[0, max_v * 1.25], dtick=max(1, round(max_v/10))),
                "yaxis": dict(gridcolor=PDF_GRID, linecolor="#cccccc", showline=True,
                              tickfont=dict(size=10)),
                "showlegend": False,
            }))
        return fig

    if not data["hires"].empty:
        if _show_rep:
            rep_fig = _pdf_grouped_bar(data["hires"], hires_past_pdf, "Staffing Rep", "#940000")
            if rep_fig is not None:
                add_chart(rep_fig, "New Hires by Staffing Rep")
        if _show_title:
            title_fig = _pdf_grouped_bar(data["hires"], hires_past_pdf, "Job Title", "#7D1F32")
            if title_fig is not None:
                add_chart(title_fig, "New Hires by Job Title")

    # ── Retention (7/30/60 Day) ──────────────────────────────────────────────
    if _show_retention and metrics["total_new_starts"] > 0:
        story.append(Paragraph("Retention (7 / 30 / 60 Day)", h2_s))
        tiers = metrics["retention_tiers"]
        total_starts = metrics["total_new_starts"]

        # Cards row — one per active tier
        ret_cards = []
        for _days in _pdf_active_days if _pdf_active_days else (7, 30, 60):
            _tier = tiers[_days]
            if _tier["pct"] is None:
                cell = _make_card_cell("N/A", f"{_days}-Day Retention", "Insufficient data")
            else:
                _sub = f"{_tier['retained']} of {_tier['eligible']} eligible"
                cell = _make_card_cell(f"{_tier['pct']:.0f}%", f"{_days}-Day Retention", _sub)
                if metrics_past is not None and metrics_past["total_new_starts"] > 0:
                    _past_tier = metrics_past["retention_tiers"][_days]
                    if _past_tier["pct"] is not None:
                        dt_text, dt_style = _pdf_delta(_tier["pct"] - _past_tier["pct"], is_pct=True, good="up")
                        cell.insert(-1, Paragraph(dt_text, dt_style))
            ret_cards.append(cell)

        n_ret_cols = len(ret_cards) if ret_cards else 1
        ret_w = 9.8 * inch / n_ret_cols
        ret_t = Table([ret_cards], colWidths=[ret_w] * n_ret_cols)
        ret_t.setStyle(TableStyle(kpi_style_cmds))
        story.append(ret_t)
        story.append(Spacer(1, 8))

        # Early term detail table — show most inclusive tier with data
        _detail_tier = next((tiers[d] for d in (60, 30, 7) if tiers[d]["pct"] is not None and not tiers[d]["early_terms"].empty), None)
        _detail_days = next((d for d in (60, 30, 7) if tiers[d]["pct"] is not None and not tiers[d]["early_terms"].empty), None)
        if _detail_tier is not None:
            et = _detail_tier["early_terms"]
            et_cols = [c for c in ["Name", "End Reason", "Days Employed", "Start Date", "End Date"] if c in et.columns]
            et_rows = [[Paragraph(c, th_s) for c in et_cols]]
            for _, row in et.iterrows():
                et_rows.append([Paragraph(safe_cell(row.get(c, "")), cell_s) for c in et_cols])
            et_widths = {"Name": 1.8*inch, "End Reason": 1.8*inch, "Days Employed": 1.0*inch,
                         "Start Date": 1.2*inch, "End Date": 1.2*inch}
            et_t = Table(et_rows, colWidths=[et_widths.get(c, 1.2*inch) for c in et_cols], repeatRows=1)
            et_t.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), SLATE),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("BOX", (0, 0), (-1, -1), 0.4, GRID),
                ("INNERGRID", (0, 0), (-1, -1), 0.4, GRID),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, STRIPE]),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]))
            story.append(Paragraph(f"Early Terminations (Within {_detail_days} Days)", h2_s))
            story.append(et_t)
            story.append(Spacer(1, 10))

    # ── Jobs Section (when present) ──────────────────────────────────────────
    if _show_jobs and "jobs" in data and not data["jobs"].empty:
        jobs_pdf = data["jobs"]
        total_openings_pdf = int(jobs_pdf["# Openings"].sum())
        total_placed_pdf   = int(jobs_pdf["Placements"].sum())
        avg_fill_pdf       = jobs_pdf["Fill Rate"].mean() * 100
        avg_days_pdf       = int(jobs_pdf["Days Opened"].mean())

        story.append(Paragraph("Jobs Overview", h2_s))
        job_cards = [
            _make_card_cell(total_openings_pdf, "Total Openings", f"across {len(jobs_pdf)} jobs"),
            _make_card_cell(total_placed_pdf, "Total Placed", f"{total_openings_pdf - total_placed_pdf} unfilled"),
            _make_card_cell(f"{avg_fill_pdf:.0f}%", "Avg Fill Rate", f"{total_placed_pdf} of {total_openings_pdf}"),
            _make_card_cell(f"{avg_days_pdf}d", "Avg Days Open", "time to fill"),
        ]
        job_card_w = 9.8 * inch / 4
        job_t = Table([job_cards], colWidths=[job_card_w] * 4)
        job_t.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "TOP"), ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("BOX", (0,0), (-1,-1), 0.5, GRID), ("INNERGRID", (0,0), (-1,-1), 0.5, GRID),
            ("TOPPADDING", (0,0), (-1,-1), 10), ("BOTTOMPADDING", (0,0), (-1,-1), 10),
            ("LEFTPADDING", (0,0), (-1,-1), 4), ("RIGHTPADDING", (0,0), (-1,-1), 4),
            ("BACKGROUND", (0,0), (-1,-1), colors.white),
            ("LINEABOVE", (0,0), (-1,0), 3, NAVY),
        ]))
        story.append(job_t)
        story.append(Spacer(1, 8))

    # ── Term tables ───────────────────────────────────────────────────────────
    def add_table(df, label):
        if df.empty:
            return
        story.append(Paragraph(label, h2_s))
        cols = [c for c in ["Name", "End Reason", "Comments", "Start Date", "End Date"] if c in df.columns]
        rows = [[Paragraph(c, th_s) for c in cols]]
        for _, row in df.iterrows():
            rows.append([Paragraph(safe_cell(row.get(c, "")), cell_s) for c in cols])
        widths = {"Name": 1.35*inch, "End Reason": 1.35*inch,
                  "Comments": 4.5*inch, "Start Date": 1.0*inch, "End Date": 1.0*inch}
        col_w = [widths.get(c, 1.2*inch) for c in cols]
        t = Table(rows, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), SLATE),
            ("FONTSIZE",      (0, 0), (-1, -1), 7.5),
            ("BOX",           (0, 0), (-1, -1), 0.4, GRID),
            ("INNERGRID",     (0, 0), (-1, -1), 0.4, GRID),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, STRIPE]),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(t)
        story.append(Spacer(1, 10))

    if _show_tables:
        add_table(metrics["involuntary"], "Involuntary Terminations (Ended by Employer)")
        add_table(metrics["voluntary"],   "Voluntary Terminations (Ended by Candidate)")
        add_table(metrics["layoffs"],     "Layoffs / Assignment Complete (Excluded from Turnover)")
        if not data["converted"].empty:
            _conv_pdf = data["converted"].copy()
            _conv_pdf_cols = [c for c in ["Name", "Job Title", "Staffing Rep", "Converted Date", "Start Date"] if c in _conv_pdf.columns]
            add_table(_conv_pdf[_conv_pdf_cols], "Conversions to Permanent")


    doc.build(story)
    buf.seek(0)
    return buf


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.image("kp_logo_white.png", width=160)
    st.markdown("---")
    st.markdown("**Current Period**")
    uploaded = st.file_uploader("Upload Current (.xlsx)", type=["xlsx"], key="upload_current")
    st.markdown("**Prior Period** *(optional)*")
    uploaded_past = st.file_uploader("Upload Past (.xlsx)", type=["xlsx"], key="upload_past")
    if uploaded is not None:
        st.caption(f"✅ Current: {uploaded.name}")
    if uploaded_past is not None:
        st.caption(f"✅ Past: {uploaded_past.name}")

import base64

# ─── Session state: clear everything when current file is removed ─────────────
if uploaded is None:
    for k in ["file_id", "parsed_data", "parsed_metrics",
              "file_id_past", "parsed_data_past", "parsed_metrics_past"]:
        st.session_state.pop(k, None)
    with open("kp_logo.png", "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
    st.markdown(f"""
    <div class="kp-header">
        <img src="data:image/png;base64,{logo_b64}" />
        <div class="kp-header-text">
            <h1>Upload a report to get started</h1>
            <p>Upload your weekly Excel report in the sidebar to get started.</p>
        </div>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ─── Parse current file ────────────────────────────────────────────────────────
import hashlib as _hashlib
_file_bytes = uploaded.getvalue()
file_id = _hashlib.md5(_file_bytes).hexdigest()
uploaded.seek(0)
if st.session_state.get("file_id") != file_id:
    try:
        data    = parse_report(uploaded)
        metrics = compute_metrics(data)
    except Exception as e:
        st.error(f"Error parsing file: {e}")
        import traceback; st.code(traceback.format_exc())
        st.stop()
    st.session_state["file_id"]        = file_id
    st.session_state["parsed_data"]    = data
    st.session_state["parsed_metrics"] = metrics

    # Auto-uncheck display settings when no data backs them
    _tiers = metrics["retention_tiers"]
    _has_terms   = not data["terms"].empty
    _has_hires   = not data["hires"].empty
    _has_hc      = not data["headcount"].empty
    _has_hours   = metrics["total_hours"] > 0
    _has_conv    = not data["converted"].empty
    _has_jobs    = not data.get("jobs", pd.DataFrame()).empty
    _has_inval   = metrics["inval_count"] > 0
    _has_vol     = metrics["vol_count"] > 0
    _has_layoff  = metrics["layoff_count"] > 0
    _has_rep     = _has_hires and "Staffing Rep" in data["hires"].columns and data["hires"]["Staffing Rep"].notna().any()
    _has_title   = _has_hires and "Job Title" in data["hires"].columns and data["hires"]["Job Title"].notna().any()
    _has_reasons = _has_terms and "End Reason" in data["terms"].columns and data["terms"]["End Reason"].str.strip().ne("").any()

    _auto = {
        "show_hc":           _has_hc,
        "show_starts":       _has_hires,
        "show_inval":        _has_inval,
        "show_vol":          _has_vol,
        "show_layoff":       _has_layoff,
        "show_turnover":     _has_inval or _has_vol,
        "show_hours":        _has_hours,
        "show_converted":    _has_conv,
        "show_hc_chart":     _has_hc,
        "show_donut":        _has_terms,
        "show_reasons":      _has_reasons,
        "show_hires":        _has_hires,
        "show_rep":          _has_rep,
        "show_title":        _has_title,
        "show_jobs":         _has_jobs,
        "show_retention_7":  _tiers[7]["pct"] is not None,
        "show_retention_30": _tiers[30]["pct"] is not None,
        "show_retention_60": _tiers[60]["pct"] is not None,
        "show_tables":       _has_terms,
    }
    for _k, _v in _auto.items():
        st.session_state[_k] = bool(_v)
else:
    data    = st.session_state["parsed_data"]
    metrics = st.session_state["parsed_metrics"]

# ─── Parse past file (optional) ───────────────────────────────────────────────
data_past    = None
metrics_past = None
if uploaded_past is not None:
    _past_bytes = uploaded_past.getvalue()
    file_id_past = _hashlib.md5(_past_bytes).hexdigest()
    uploaded_past.seek(0)
    if st.session_state.get("file_id_past") != file_id_past:
        try:
            data_past    = parse_report(uploaded_past)
            metrics_past = compute_metrics(data_past)
        except Exception as e:
            st.warning(f"Could not parse prior period file: {e}")
            data_past = metrics_past = None
        if data_past is not None:
            st.session_state["file_id_past"]        = file_id_past
            st.session_state["parsed_data_past"]    = data_past
            st.session_state["parsed_metrics_past"] = metrics_past
    else:
        data_past    = st.session_state.get("parsed_data_past")
        metrics_past = st.session_state.get("parsed_metrics_past")
else:
    for k in ["file_id_past", "parsed_data_past", "parsed_metrics_past"]:
        st.session_state.pop(k, None)

# ─── Date Range Filters ───────────────────────────────────────────────────────
def _date_bounds(parsed_data):
    """Return (min_date, max_date) across all data in the parsed report."""
    all_dates = []
    hc = parsed_data.get("headcount", pd.DataFrame())
    if not hc.empty and "Week Ending" in hc.columns:
        all_dates += hc["Week Ending"].dropna().dt.date.tolist()
    terms = parsed_data.get("terms", pd.DataFrame())
    if not terms.empty and "End Date" in terms.columns:
        all_dates += terms["End Date"].dropna().dt.date.tolist()
    hires = parsed_data.get("hires", pd.DataFrame())
    if not hires.empty and "Start Date" in hires.columns:
        all_dates += hires["Start Date"].dropna().dt.date.tolist()
    if not all_dates:
        return None, None
    return min(all_dates), max(all_dates)

_cur_min, _cur_max = _date_bounds(data)
_past_min, _past_max = _date_bounds(data_past) if data_past is not None else (None, None)

with st.sidebar:
    if _cur_min is not None:
        st.markdown("---")
        st.markdown("**📅 Date Filter — Current**")
        cur_filter_start = st.date_input(
            "From", value=_cur_min,
            min_value=_cur_min, max_value=_cur_max,
            key="date_filter_cur_start",
        )
        cur_filter_end = st.date_input(
            "To", value=_cur_max,
            min_value=_cur_min, max_value=_cur_max,
            key="date_filter_cur_end",
        )
        # Guard: if user sets end before start, swap
        if cur_filter_end < cur_filter_start:
            cur_filter_start, cur_filter_end = cur_filter_end, cur_filter_start
    else:
        cur_filter_start, cur_filter_end = None, None

    if _past_min is not None:
        st.markdown("**📅 Date Filter — Prior**")
        past_filter_start = st.date_input(
            "From", value=_past_min,
            min_value=_past_min, max_value=_past_max,
            key="date_filter_past_start",
        )
        past_filter_end = st.date_input(
            "To", value=_past_max,
            min_value=_past_min, max_value=_past_max,
            key="date_filter_past_end",
        )
        if past_filter_end < past_filter_start:
            past_filter_start, past_filter_end = past_filter_end, past_filter_start
    else:
        past_filter_start, past_filter_end = None, None

    st.markdown("---")
    st.markdown("**Display Settings**")
    show_headcount_card = st.checkbox("Starting Headcount", value=True, key="show_hc")
    show_starts_card    = st.checkbox("Total Starts", value=True, key="show_starts")
    show_inval_card     = st.checkbox("Involuntary Terms", value=True, key="show_inval")
    show_vol_card       = st.checkbox("Voluntary Terms", value=True, key="show_vol")
    show_layoff_card    = st.checkbox("Layoffs / Assign. Complete", value=True, key="show_layoff")
    show_turnover_card  = st.checkbox("Total Turnover", value=True, key="show_turnover")
    show_hours          = st.checkbox("Total Hours", value=True, key="show_hours")
    show_converted      = st.checkbox("Converted Employees", value=True, key="show_converted")
    st.markdown("---")
    st.markdown("**Charts**")
    show_headcount_chart = st.checkbox("Headcount by Week", value=True, key="show_hc_chart")
    show_donut           = st.checkbox("Terminations by Type", value=True, key="show_donut")
    show_reasons         = st.checkbox("Termination Reasons", value=True, key="show_reasons")
    show_hires_chart     = st.checkbox("New Hires by Week", value=True, key="show_hires")
    show_rep_chart       = st.checkbox("Hires by Staffing Rep", value=True, key="show_rep")
    show_title_chart     = st.checkbox("Hires by Job Title", value=True, key="show_title")
    show_jobs_section    = st.checkbox("Jobs Overview", value=True, key="show_jobs")
    st.markdown("**Retention**")
    show_retention_7     = st.checkbox("7-Day Retention",  value=True, key="show_retention_7")
    show_retention_30    = st.checkbox("30-Day Retention", value=True, key="show_retention_30")
    show_retention_60    = st.checkbox("60-Day Retention", value=True, key="show_retention_60")
    show_retention       = show_retention_7 or show_retention_30 or show_retention_60
    show_term_tables     = st.checkbox("Termination Detail Tables", value=True, key="show_tables")
    st.markdown("---")
    st.markdown("KP Staffing · Activity Report")

# Apply date filters
if cur_filter_start is not None:
    data_filtered = filter_data_by_dates(data, cur_filter_start, cur_filter_end)
    metrics = compute_metrics(data_filtered)
else:
    data_filtered = data

if data_past is not None and past_filter_start is not None:
    data_past_filtered = filter_data_by_dates(data_past, past_filter_start, past_filter_end)
    metrics_past = compute_metrics(data_past_filtered)
else:
    data_past_filtered = data_past

# ─── Header ───────────────────────────────────────────────────────────────────
with open("kp_logo.png", "rb") as f:
    logo_b64 = base64.b64encode(f.read()).decode()
st.markdown(f"""
<div class="kp-header">
    <img src="data:image/png;base64,{logo_b64}" />
    <div class="kp-header-text">
        <h1>{data['company']}</h1>
        <p>Period: {cur_filter_start.strftime('%b %d, %Y') + ' – ' + cur_filter_end.strftime('%b %d, %Y') if (cur_filter_start is not None and (cur_filter_start != _cur_min or cur_filter_end != _cur_max)) else data['date_range']}</p>
    </div>
</div>
""", unsafe_allow_html=True)

# ─── KPI Cards ────────────────────────────────────────────────────────────────

def _delta_html(diff, good="down", is_pct=False):
    """Return a styled delta badge. Colors flip based on 'good' direction.
    good='up'   → increase=green, decrease=red  (headcount, starts, hours)
    good='down' → increase=red,   decrease=green (terms, turnover)
    good='neutral' → always neutral color
    """
    suffix = "pp" if is_pct else ""
    if abs(diff) < (0.05 if is_pct else 0.5):
        return '<div class="delta-neutral">→ No change vs prior</div>'
    arrow = "↑" if diff > 0 else "↓"
    sign  = "+" if diff > 0 else ""
    val   = f"{sign}{diff:.1f}{suffix}" if is_pct else f"{sign}{int(diff)}"
    if good == "neutral":
        css = "delta-neutral"
    elif good == "up":
        css = "delta-good" if diff > 0 else "delta-bad"
    else:  # good == "down"
        css = "delta-bad" if diff > 0 else "delta-good"
    return f'<div class="{css}">{arrow} <strong>{val}</strong> vs prior</div>'

def simple_card(col, val, label, sub, delta_html=""):
    with col:
        parts = [
            '<div class="metric-card">',
            f'<div class="metric-value">{val}</div>',
            f'<div class="metric-label">{label}</div>',
        ]
        if delta_html:
            parts.append(delta_html)
        parts += [f'<div class="metric-sub">{sub}</div>', '</div>']
        st.markdown("".join(parts), unsafe_allow_html=True)

def pct_card(col, val, label, sub, pct, delta_count_html="", delta_pct_html=""):
    with col:
        parts = [
            '<div class="metric-card">',
            f'<div class="metric-value">{val}</div>',
            f'<div class="metric-label">{label}</div>',
        ]
        if delta_count_html:
            parts.append(delta_count_html)
        parts += [
            '<div class="metric-divider"></div>',
            f'<div class="metric-pct">{pct}</div>',
        ]
        if delta_pct_html:
            parts.append(delta_pct_html)
        parts += [f'<div class="metric-sub">{sub}</div>', '</div>']
        st.markdown("".join(parts), unsafe_allow_html=True)

# ─── Comparison banner + toggle ───────────────────────────────────────────────
view_mode = "current"  # default
if metrics_past is not None:
    banner_col, toggle_col = st.columns([5, 1])
    with banner_col:
        st.success(f"📊 Comparing **{data['date_range']}** vs prior period **{data_past['date_range']}**")
    with toggle_col:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        view_mode = st.radio("View", ["Current", "Prior"], index=0, horizontal=True, label_visibility="collapsed")
        view_mode = view_mode.lower()
elif uploaded_past is not None:
    st.warning("⚠️ Past file detected but could not be parsed — check it's a valid activity report.")

# When viewing prior period, swap data/metrics to show past numbers (no deltas)
if view_mode == "prior" and data_past_filtered is not None and metrics_past is not None:
    _active_data    = data_past_filtered
    _active_metrics = metrics_past
    _show_deltas    = False
else:
    _active_data    = data_filtered
    _active_metrics = metrics
    _show_deltas    = (metrics_past is not None)

# ─── Pre-compute deltas (only when showing current with comparison) ────────────
d: dict = {}  # holds delta html strings — only populated when showing deltas
if _show_deltas:
    p = metrics_past
    d["hc"]          = _delta_html(
        (metrics["starting_headcount"] or metrics["adj_placements"]) -
        (p["starting_headcount"] or p["adj_placements"]), good="up")
    d["starts"]      = _delta_html(data["new_placements_raw"] - data_past["new_placements_raw"], good="up")
    d["inval_cnt"]   = _delta_html(metrics["inval_count"]  - p["inval_count"],  good="down")
    d["inval_pct"]   = _delta_html(metrics["inval_pct"]    - p["inval_pct"],    good="down", is_pct=True)
    d["vol_cnt"]     = _delta_html(metrics["vol_count"]    - p["vol_count"],    good="down")
    d["vol_pct"]     = _delta_html(metrics["vol_pct"]      - p["vol_pct"],      good="down", is_pct=True)
    d["layoff"]      = _delta_html(metrics["layoff_count"] - p["layoff_count"], good="neutral")
    d["turnover_pct"]= _delta_html(metrics["total_turnover_pct"] - p["total_turnover_pct"], good="down", is_pct=True)
    d["hours"]       = _delta_html(metrics["total_hours"] - p["total_hours"], good="up")

m = _active_metrics
ad = _active_data
converted_count = len(ad["converted"]) if not ad["converted"].empty else 0
total_hours_val = f"{int(m['total_hours']):,}" if m["total_hours"] > 0 else "0"

# ─── Build card groups ────────────────────────────────────────────────────────

def _render_card(col, cd):
    if cd[0] == "simple":
        simple_card(col, cd[1], cd[2], cd[3], delta_html=cd[4])
    else:
        pct_card(col, cd[1], cd[2], cd[3], cd[4], delta_count_html=cd[5], delta_pct_html=cd[6])

def _render_row(cards):
    if not cards:
        return
    cols = st.columns(len(cards))
    for i, cd in enumerate(cards):
        _render_card(cols[i], cd)
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

# ── Row 1: Workforce Overview ─────────────────────────────────────────────────
_row1 = []
if show_headcount_card:
    _row1.append(("simple", m["starting_headcount"] or m["adj_placements"],
                  "Starting Headcount", "headcount at period start", d.get("hc", "")))
if show_starts_card:
    _row1.append(("simple", ad.get("new_placements_raw", 0), "Total Starts",
                  "new placements this period", d.get("starts", "")))
if show_hours:
    _row1.append(("simple", total_hours_val, "Total Hours", "billed this period",
                  d.get("hours", "")))
_render_row(_row1)

# ── Row 2: Turnover ───────────────────────────────────────────────────────────
_row2 = []
if show_inval_card:
    _row2.append(("pct", m["inval_count"], "Involuntary Terms", "of total active employees",
                  f"{m['inval_pct']:.1f}%", d.get("inval_cnt", ""), d.get("inval_pct", "")))
if show_vol_card:
    _row2.append(("pct", m["vol_count"], "Voluntary Terms", "of total active employees",
                  f"{m['vol_pct']:.1f}%", d.get("vol_cnt", ""), d.get("vol_pct", "")))
if show_turnover_card:
    _row2.append(("pct", f"{m['total_turnover_pct']:.1f}%", "Total Turnover",
                  "excl. layoffs and conversions",
                  f"{m['inval_count'] + m['vol_count']} total terms",
                  "", d.get("turnover_pct", "")))
_render_row(_row2)

# ── Row 3: Retention ──────────────────────────────────────────────────────────
_row3 = []
if m["total_new_starts"] > 0:
    for _days, _label, _show in [(7, "7-Day", show_retention_7), (30, "30-Day", show_retention_30), (60, "60-Day", show_retention_60)]:
        if not _show:
            continue
        _tier = m["retention_tiers"][_days]
        _ret_delta = ""
        if _tier["pct"] is None:
            _row3.append(("simple", "N/A", f"{_label} Retention", "Insufficient data", ""))
            continue
        if _show_deltas and metrics_past is not None and metrics_past["total_new_starts"] > 0:
            _past_tier = metrics_past["retention_tiers"][_days]
            if _past_tier["pct"] is not None:
                _ret_delta = _delta_html(_tier["pct"] - _past_tier["pct"], good="up", is_pct=True)
        _row3.append(("simple", f"{_tier['pct']:.0f}%", f"{_label} Retention",
                      f"{_tier['retained']} of {_tier['eligible']} eligible", _ret_delta))
_render_row(_row3)

# ── Row 4: Exclusions & Conversions ──────────────────────────────────────────
_row4 = []
if show_layoff_card:
    _row4.append(("simple", m["layoff_count"], "Layoffs / Assignment Complete",
                  "excluded from turnover", d.get("layoff", "")))
if show_converted and converted_count > 0:
    _row4.append(("simple", converted_count, "Converted Employees", "hired by client this period", ""))
_render_row(_row4)

st.markdown("<br>", unsafe_allow_html=True)

# ─── Row 1: Headcount + Donut ─────────────────────────────────────────────────
_show_hc = show_headcount_chart and not ad["headcount"].empty
_show_dn = show_donut
if _show_hc and _show_dn:
    col1, col2 = st.columns([3, 2])
    with col1:
        st.markdown('<div class="section-header">Headcount by Week</div>', unsafe_allow_html=True)
        st.plotly_chart(make_fig_headcount(ad["headcount"], include_hours=show_hours), use_container_width=True)
    with col2:
        st.markdown('<div class="section-header">Terminations by Type</div>', unsafe_allow_html=True)
        st.plotly_chart(make_fig_donut(m), use_container_width=True)
elif _show_hc:
    st.markdown('<div class="section-header">Headcount by Week</div>', unsafe_allow_html=True)
    st.plotly_chart(make_fig_headcount(ad["headcount"], include_hours=show_hours), use_container_width=True)
elif _show_dn:
    st.markdown('<div class="section-header">Terminations by Type</div>', unsafe_allow_html=True)
    st.plotly_chart(make_fig_donut(m), use_container_width=True)

# ─── Row 2: Reasons ───────────────────────────────────────────────────────────
if show_reasons:
    # In current+comparison mode, show grouped bars; in prior-only mode, show just prior data
    if _show_deltas:
        _past_inval = metrics_past["involuntary"]
        _past_vol   = metrics_past["voluntary"]
    else:
        _past_inval = None
        _past_vol   = None

    col3, col4 = st.columns(2)
    with col3:
        st.markdown('<div class="section-header">Involuntary Termination Reasons</div>', unsafe_allow_html=True)
        st.plotly_chart(make_fig_reasons(m["involuntary"],
            ["#940000","#7D1F32","#B85050","#C0392B","#5C0A0A","#A03030","#D4A0A0","#6B1515"],
            df_past=_past_inval), use_container_width=True)
    with col4:
        st.markdown('<div class="section-header">Voluntary Termination Reasons</div>', unsafe_allow_html=True)
        st.plotly_chart(make_fig_reasons(m["voluntary"],
            ["#2C3E50","#34495E","#4A6278","#5D7A8A","#1A252F","#3D566E","#6B8CA0","#8AA0B0"],
            df_past=_past_vol), use_container_width=True)

# ─── Row 3: New Hires by Start Week ──────────────────────────────────────────
_past_hires = data_past["hires"] if _show_deltas and data_past is not None else None
if show_hires_chart:
    st.markdown('<div class="section-header">New Hires by Start Week</div>', unsafe_allow_html=True)
    st.plotly_chart(make_fig_hires(ad["hires"]), use_container_width=True)

# ─── Row 4: Staffing Rep + Job Title ─────────────────────────────────────────
if show_rep_chart or show_title_chart:
    rep_fig   = make_fig_rep_hires(ad["hires"],  hires_past=_past_hires) if show_rep_chart else None
    title_fig = make_fig_job_title_hires(ad["hires"], hires_past=_past_hires) if show_title_chart else None
    if show_rep_chart and show_title_chart and rep_fig and rep_fig.data and title_fig and title_fig.data:
        hire_col1, hire_col2 = st.columns(2)
        with hire_col1:
            st.markdown('<div class="section-header">New Hires by Staffing Rep</div>', unsafe_allow_html=True)
            st.plotly_chart(rep_fig, use_container_width=True)
        with hire_col2:
            st.markdown('<div class="section-header">New Hires by Job Title</div>', unsafe_allow_html=True)
            st.plotly_chart(title_fig, use_container_width=True)
    elif show_rep_chart and rep_fig and rep_fig.data:
        st.markdown('<div class="section-header">New Hires by Staffing Rep</div>', unsafe_allow_html=True)
        st.plotly_chart(rep_fig, use_container_width=True)
    elif show_title_chart and title_fig and title_fig.data:
        st.markdown('<div class="section-header">New Hires by Job Title</div>', unsafe_allow_html=True)
        st.plotly_chart(title_fig, use_container_width=True)

# ─── Retention (7/30/60 Day) ──────────────────────────────────────────────────
_active_retention_tiers = [(d, l, s) for d, l, s in [(7, "7-Day", show_retention_7), (30, "30-Day", show_retention_30), (60, "60-Day", show_retention_60)] if s]
if _active_retention_tiers and m["total_new_starts"] > 0:
    _tier_labels = " / ".join(l for _, l, _ in _active_retention_tiers)
    st.markdown(f'<div class="section-header">Retention ({_tier_labels})</div>', unsafe_allow_html=True)
    _mp = metrics_past if _show_deltas and metrics_past is not None and metrics_past["total_new_starts"] > 0 else None
    _active_days = [d for d, _, _ in _active_retention_tiers]
    st.plotly_chart(make_fig_retention_tiers(m, _mp, active_days=_active_days), use_container_width=True)

    def _style_retention_table(df):
        return df.style.set_properties(**{
            "background-color": "white", "color": "#131313", "font-size": "13px",
        }).set_table_styles([
            {"selector": "thead th", "props": [
                ("background-color", "#940000"), ("color", "white"),
                ("font-weight", "600"), ("font-size", "13px"),
            ]},
            {"selector": "tbody tr:nth-child(even) td", "props": [("background-color", "#fdf5f5")]},
            {"selector": "tbody tr:nth-child(odd) td",  "props": [("background-color", "white")]},
        ])

    # Show early termination details for each active tier
    for _days, _, _ in _active_retention_tiers:
        _tier = m["retention_tiers"][_days]
        if not _tier["early_terms"].empty:
            with st.expander(f"Left within {_days} days — {_tier['early_count']} employees"):
                et_disp = _tier["early_terms"].copy()
                et_cols = [c for c in ["Name", "End Reason", "Days Employed", "Start Date", "End Date"] if c in et_disp.columns]
                for col in ["Start Date", "End Date"]:
                    if col in et_disp.columns:
                        et_disp[col] = pd.to_datetime(et_disp[col], errors="coerce").dt.strftime("%b %d, %Y").fillna("")
                st.dataframe(_style_retention_table(et_disp[et_cols]), use_container_width=True, hide_index=True)

# ─── Jobs Section (when present) ─────────────────────────────────────────────
if show_jobs_section and "jobs" in ad and not ad["jobs"].empty:
    jobs = ad["jobs"]
    total_openings = int(jobs["# Openings"].sum())
    total_placed   = int(jobs["Placements"].sum())
    avg_fill       = jobs["Fill Rate"].mean() * 100
    avg_days       = int(jobs["Days Opened"].mean())

    st.markdown('<div class="section-header">Jobs Overview</div>', unsafe_allow_html=True)

    jc1, jc2, jc3, jc4 = st.columns(4)
    for col, val, label, sub in [
        (jc1, total_openings, "Total Openings",   f"across {len(jobs)} jobs"),
        (jc2, total_placed,   "Total Placed",      f"{total_openings - total_placed} unfilled"),
        (jc3, f"{avg_fill:.0f}%", "Avg Fill Rate", f"{total_placed} of {total_openings}"),
        (jc4, f"{avg_days}d",  "Avg Days Open",    "time to fill"),
    ]:
        simple_card(col, val, label, sub)

    st.markdown("<br>", unsafe_allow_html=True)

    jcol1, jcol2 = st.columns(2)
    with jcol1:
        st.markdown('<div class="section-header">Openings vs Placed by Job</div>', unsafe_allow_html=True)
        st.plotly_chart(make_fig_jobs_fill(jobs), use_container_width=True)
    with jcol2:
        st.markdown('<div class="section-header">Days Open by Job</div>', unsafe_allow_html=True)
        st.plotly_chart(make_fig_jobs_days(jobs), use_container_width=True)

    st.markdown('<div class="section-header">Fill Rate by Job</div>', unsafe_allow_html=True)
    st.plotly_chart(make_fig_jobs_fillrate(jobs), use_container_width=True)

    with st.expander(f"Jobs Detail Table ({len(jobs)} jobs)"):
        disp = jobs.copy()
        for col in ["Date Added", "Date Closed"]:
            if col in disp.columns:
                disp[col] = disp[col].dt.strftime("%b %d, %Y").fillna("")
        disp["Fill Rate"] = (disp["Fill Rate"] * 100).round(1).astype(str) + "%"
        st.dataframe(disp, use_container_width=True, hide_index=True)

# ─── Termination Tables ───────────────────────────────────────────────────────
if show_term_tables:
    st.markdown('<div class="section-header">Termination Detail</div>', unsafe_allow_html=True)
    _conv_count = len(ad["converted"]) if not ad["converted"].empty else 0
    tab1, tab2, tab3, tab4 = st.tabs([
        f"🔴 Involuntary ({m['inval_count']})",
        f"🔵 Voluntary ({m['vol_count']})",
        f"🟡 Layoffs / Assignment Complete ({m['layoff_count']})",
        f"🟢 Conversions ({_conv_count})",
    ])
    def style_table(df):
        return df.style.set_properties(**{
            "background-color": "white",
            "color": "#131313",
            "font-size": "13px",
        }).set_table_styles([
            {"selector": "thead th", "props": [
                ("background-color", "#940000"),
                ("color", "white"),
                ("font-weight", "600"),
                ("font-size", "13px"),
            ]},
            {"selector": "tbody tr:nth-child(even) td", "props": [("background-color", "#fdf5f5")]},
            {"selector": "tbody tr:nth-child(odd) td",  "props": [("background-color", "white")]},
        ])

    with tab1:
        st.dataframe(style_table(format_terms_display(m["involuntary"])), use_container_width=True, hide_index=True)
    with tab2:
        st.dataframe(style_table(format_terms_display(m["voluntary"])), use_container_width=True, hide_index=True)
    with tab3:
        st.dataframe(style_table(format_terms_display(m["layoffs"])), use_container_width=True, hide_index=True)
    with tab4:
        if ad["converted"].empty:
            st.info("No conversions recorded for this period.")
        else:
            _conv_disp = ad["converted"].copy()
            _conv_cols = [c for c in ["Name", "Job Title", "Staffing Rep", "Converted Date", "Start Date"] if c in _conv_disp.columns]
            for _dc in ["Converted Date", "Start Date"]:
                if _dc in _conv_disp.columns:
                    _conv_disp[_dc] = pd.to_datetime(_conv_disp[_dc], errors="coerce").dt.strftime("%m/%d/%Y")
            st.dataframe(style_table(_conv_disp[_conv_cols]), use_container_width=True, hide_index=True)


# ─── Export ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown("### 📄 Export Report")
_export_col1, _export_col2 = st.columns([1, 1])

with _export_col2:
    _xl_fname = f"{_active_data['company'].replace(' ', '_')}_staffing_report.xlsx"
    with st.spinner(""):
        _xl_buf = generate_excel(_active_data, _active_metrics,
                                 data_past=(data_past if view_mode == "current" else None),
                                 metrics_past=(metrics_past if view_mode == "current" else None))
    st.download_button("⬇️ Download Excel", data=_xl_buf, file_name=_xl_fname,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with _export_col1:
    if st.button("Generate PDF Report"):
        with st.spinner("Building PDF..."):
            _pdf_opts = {
                "show_hours":    show_hours,
                "show_converted":show_converted,
                "show_hc_chart": show_headcount_chart,
                "show_donut":    show_donut,
                "show_reasons":  show_reasons,
                "show_hires":    show_hires_chart,
                "show_rep":      show_rep_chart,
                "show_title":    show_title_chart,
                "show_retention":   show_retention,
                "show_retention_7": show_retention_7,
                "show_retention_30":show_retention_30,
                "show_retention_60":show_retention_60,
                "show_jobs":     show_jobs_section,
                "show_tables":   show_term_tables,
                "period_label":  (f"{cur_filter_start.strftime('%b %d, %Y')} – {cur_filter_end.strftime('%b %d, %Y')}"
                                  if cur_filter_start and (cur_filter_start != _cur_min or cur_filter_end != _cur_max)
                                  else _active_data.get("date_range", "")),
                "past_period_label": (f"{past_filter_start.strftime('%b %d, %Y')} – {past_filter_end.strftime('%b %d, %Y')}"
                                      if past_filter_start and data_past and (past_filter_start != _past_min or past_filter_end != _past_max)
                                      else (data_past or {}).get("date_range", "")),
            }
            _pdf_past_data    = data_past    if view_mode == "current" else None
            _pdf_past_metrics = metrics_past if view_mode == "current" else None
            pdf_buf = generate_pdf(
                _active_data, _active_metrics,
                data_past=_pdf_past_data, metrics_past=_pdf_past_metrics,
                display_opts=_pdf_opts,
            )
        fname = f"{_active_data['company'].replace(' ', '_')}_staffing_report.pdf"
        st.download_button("⬇️ Download PDF", data=pdf_buf, file_name=fname, mime="application/pdf")
